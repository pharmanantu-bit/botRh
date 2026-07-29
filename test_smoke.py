"""Test de fumée — charge chaque page de l'app avec des données réalistes et
vérifie qu'aucune ne plante (HTTP 200). Lancé en CI à chaque push : aurait
attrapé les bugs dashboard/absences avant le déploiement.

Usage : python test_smoke.py   (code de sortie non nul si une page échoue)
"""
import os
import re
import sys
import json
import urllib.parse
from datetime import datetime

import app as A
import tokens
import agent_rh
import crypto_rh
import extraction_pj

A.app.config["TESTING"] = True
client = A.app.test_client()

mois = datetime.now().month
annee = datetime.now().year
employes = A.charger_employes()

# Donnée de test : une réponse pour le 1er employé, pour exercer les recherches
fichier_temp = A.reponses_file(mois, annee)
cree_temp = not os.path.exists(fichier_temp)
if employes and cree_temp:
    e = employes[0]
    tok = tokens.generer_token(e["prenom"], e["email"])
    with open(fichier_temp, "w", encoding="utf-8") as f:
        json.dump({tok: {"prenom": e["prenom"], "heures_plus": 6.0, "heures_moins": 1.5,
                         "commentaire": "smoke", "date_signature": "10/06/2026",
                         "signature": "Test", "date": "10/06/2026 10:00",
                         "mois": mois, "annee": annee}}, f, ensure_ascii=False)

cle = A.API_CLE
prenom0 = employes[0]["prenom"] if employes else "Test"
tok0 = tokens.generer_token(employes[0]["prenom"], employes[0]["email"]) if employes else "x"

routes_publiques = [
    (f"/releve?token={tok0}&prenom={prenom0}", 200),
    (f"/mon-espace?token={tok0}&prenom={prenom0}", 200),
    (f"/healthcheck?cle={cle}", 200),
    (f"/export_reponses?cle={cle}&mois={mois}&annee={annee}", 200),
    (f"/export_employes?cle={cle}", 200),
    (f"/export_recap?cle={cle}&mois={mois}&annee={annee}", 200),
    (f"/export_backup?cle={cle}", 200),
    ("/export_reponses?cle=mauvaise", 403),
    ("/docs-embauche/fiche-signaletique", 200),
    ("/docs-embauche/mutuelle", 200),
    ("/docs-embauche/inconnu", 404),
]
routes_admin = [
    "/admin/dashboard", "/admin/mois", f"/admin/mois?mois={mois}&annee={annee}",
    "/admin/historique", f"/admin/historique/{mois}/{annee}",
    "/admin/employes", "/admin/export", "/admin/erreurs",
    "/admin/sauvegarde", "/admin/sauvegarde/telecharger", "/admin/assistant", "/admin/securite",
    "/admin/recrutement",
    "/admin/planning-equipe",
    "/admin/planning-equipe?onglet=trame",
    "/admin/planning-equipe?onglet=trame&sem=B",
    "/admin/planning-equipe?onglet=effectifs",
    "/admin/planning-equipe?onglet=changements",
    "/admin/planning-equipe?onglet=conges",
    f"/admin/planning-equipe?onglet=conges&periode={annee}",
    "/admin/planning-equipe?onglet=effectifs&date=2026-07-13",
    "/admin/planning-equipe?onglet=totaux",
    f"/admin/planning-equipe?onglet=totaux&mois={annee}-{int(mois):02d}",
    "/admin/planning-equipe?onglet=totaux&mois=n_importe_quoi",
]
if employes:
    routes_admin.append(f"/admin/employe?email={urllib.parse.quote(employes[0]['email'])}")

# Routes qui redirigent (302) : /admin -> dashboard, absences/synthese fusionnées
routes_admin_redirect = ["/admin", "/admin/absences", "/admin/synthese"]

echecs = []

for url, attendu in routes_publiques:
    code = client.get(url).status_code
    ok = code == attendu
    print(f"{'OK ' if ok else 'KO '}[{code}] {url}")
    if not ok:
        echecs.append(f"{url} (attendu {attendu}, reçu {code})")

with client.session_transaction() as s:
    s["admin"] = True
for url in routes_admin:
    code = client.get(url).status_code
    ok = code == 200
    print(f"{'OK ' if ok else 'KO '}[{code}] {url}")
    if not ok:
        echecs.append(f"{url} (attendu 200, reçu {code})")

for url in routes_admin_redirect:
    code = client.get(url).status_code
    ok = code in (301, 302)
    print(f"{'OK ' if ok else 'KO '}[{code}] {url} (redirection attendue)")
    if not ok:
        echecs.append(f"{url} (attendu 302, reçu {code})")

# --- Cycle de paie : clôture au 25, période 26→25, Excel « Validé », rappel validation ---
if A.JOUR_CLOTURE == 25:
    print("OK [--] clôture : JOUR_CLOTURE = 25")
else:
    echecs.append(f"JOUR_CLOTURE = {A.JOUR_CLOTURE} (attendu 25 — .env local surchargé ?)")
    print(f"KO [--] JOUR_CLOTURE = {A.JOUR_CLOTURE}")

# Grille du formulaire : du 26 du mois précédent au 25 du mois courant inclus.
if employes and datetime.now().day <= A.JOUR_CLOTURE:
    html_releve = client.get(f"/releve?token={tok0}&prenom={prenom0}").data.decode("utf-8")
    attendus = ['name="prec_plus_26"', 'name="mois_plus_25"']
    interdits = ['name="prec_plus_24"', 'name="prec_plus_25"', 'name="mois_plus_26"', 'name="mois_plus_31"']
    grille_ok = (all(a in html_releve for a in attendus)
                 and not any(i in html_releve for i in interdits))
    if grille_ok:
        print("OK [--] grille /releve : période 26 (mois préc.) - 25 (mois courant)")
    else:
        echecs.append("grille /releve : bornes de période incorrectes")
        print("KO [--] grille /releve : bornes de période incorrectes")

# extraire_detail_jours ignore les champs hors période (24-25 préc., 26+ courant).
detail = A.extraire_detail_jours(
    {"prec_plus_24": "5", "prec_plus_26": "2", "mois_plus_25": "1", "mois_plus_26": "3"},
    mois, annee)
labels = [d["label"] for d in detail]
if len(detail) == 2 and any("26/" in l for l in labels) and any("25/" in l for l in labels):
    print("OK [--] extraire_detail_jours : champs hors période ignorés")
else:
    echecs.append(f"extraire_detail_jours : attendu 2 jours (26 préc. + 25 courant), reçu {labels}")
    print(f"KO [--] extraire_detail_jours : {labels}")

# periode_paie : la référence unique planning/relevé pour chaque mois.
from datetime import date as _d
cas_periodes = [
    ((6, 2026), (_d(2026, 5, 24), _d(2026, 6, 30))),    # historique 24 -> fin
    ((7, 2026), (_d(2026, 6, 24), _d(2026, 7, 31))),    # fin juin + tout juillet (voulu)
    ((8, 2026), (_d(2026, 8, 1), _d(2026, 8, 25))),     # transition
    ((9, 2026), (_d(2026, 8, 26), _d(2026, 9, 25))),    # régime de croisière
    ((1, 2027), (_d(2026, 12, 26), _d(2027, 1, 25))),   # passage d'année
]
rates_p = [(m, a, A.periode_paie(m, a)) for (m, a), attendu in cas_periodes
           if A.periode_paie(m, a) != attendu]
if not rates_p:
    print("OK [--] periode_paie : historique / transition août / croisière / passage d'année")
else:
    echecs.append(f"periode_paie : {rates_p}")
    print(f"KO [--] periode_paie : {rates_p}")

# Transition août 2026 : relevé du 1er au 25 (les jours de juillet, déjà
# couverts par l'ancienne grille de juillet, ne sont plus proposés ni lus).
d_aout = A.extraire_detail_jours({"prec_plus_28": "5", "mois_plus_10": "2"}, 8, 2026)
if (A.sans_jours_prec(8, 2026) and not A.sans_jours_prec(9, 2026)
        and len(d_aout) == 1 and "10/08" in d_aout[0]["label"]):
    print("OK [--] transition août 2026 : 1er-25, jours de juillet ignorés, régime normal en sept.")
else:
    echecs.append(f"transition août 2026 : {d_aout}")
    print(f"KO [--] transition août 2026 : {d_aout}")

# Excel paie : colonne « Validé » en 10e position, titre fusionné A1:J1.
from openpyxl import load_workbook
wb = load_workbook(A.construire_recap_xlsx(mois, annee))
ws = wb.active
entetes = [c.value for c in ws[2]]
fusion_ok = "A1:J1" in [str(r) for r in ws.merged_cells.ranges]
if len(entetes) == 10 and entetes[-1] == "Validé" and fusion_ok:
    print("OK [--] récap Excel : 10 colonnes dont « Validé », titre A1:J1")
else:
    echecs.append(f"récap Excel : en-têtes={entetes}, fusion A1:J1={fusion_ok}")
    print("KO [--] récap Excel : colonne Validé / fusion incorrecte")

# Mail de rappel de validation (constructeur pur, sans SMTP).
import validation_sender
_emps = [{"prenom": "Alice", "nom": "Un", "email": "a@t.fr"},
         {"prenom": "Bob", "nom": "Deux", "email": "b@t.fr"},
         {"prenom": "Chloe", "nom": "Trois", "email": "c@t.fr"}]
_reps = {
    tokens.generer_token("Alice", "a@t.fr"): {"prenom": "Alice", "valide": True,
                                              "date": "20/07/2026 10:00"},
    tokens.generer_token("Bob", "b@t.fr"): {"prenom": "Bob", "date": "21/07/2026 11:00"},
}
_sujet, _corps = validation_sender.construire_mail_rappel(_emps, _reps, "Juillet 2026")
rappel_ok = ("1 relevé(s) non validé(s)" in _sujet and "1 manquant(s)" in _sujet
             and "Bob Deux" in _corps and "Chloe Trois" in _corps
             and "Alice" not in _corps and "/admin/mois" in _corps)
_sujet_ok, _corps_ok = validation_sender.construire_mail_rappel(
    _emps[:1], {k: v for k, v in _reps.items() if v.get("valide")}, "Juillet 2026")
rappel_ok = rappel_ok and "tout est validé" in _sujet_ok
if rappel_ok:
    print("OK [--] rappel validation : sections non validés/manquants + cas tout validé")
else:
    echecs.append(f"rappel validation : sujet={_sujet!r}")
    print("KO [--] rappel validation : contenu incorrect")

# --- Agent RH outillé : chaîne complète hors-ligne (moteur fake, coût nul) ---
# Vérifie : qu'un outil est déclenché, que le prénom tapé par l'utilisateur est
# pseudonymisé (étiquette « Employé X ») AVANT d'atteindre l'outil — donc le modèle
# ne verrait jamais le vrai nom — et que la réponse finale est ré-identifiée.
if employes:
    captures = []
    def _executer_spy(nom, args, annuaire):
        captures.append((nom, dict(args or {})))
        return A.executer_outil_agent(nom, args, annuaire)

    res = agent_rh.run_agent(
        [{"role": "user", "content": f"la fiche de {prenom0}"}],
        employes, _executer_spy, moteur="fake")

    if not res.get("outils_utilises"):
        echecs.append("agent fake : aucun outil déclenché")
        print("KO [--] agent fake : aucun outil déclenché")
    else:
        ok_label = True
        for nom, args in captures:
            if nom == "profil_salarie":
                emp = args.get("employe", "")
                if not re.match(r"Employé [A-Z]+$", emp) or prenom0.lower() in emp.lower():
                    ok_label = False
                    echecs.append(f"agent fake : nom non pseudonymisé dans l'outil ({emp!r})")
        # La réponse doit être ré-identifiée (le prénom réel réapparaît à l'affichage).
        ok_reid = prenom0 in res.get("reply", "")
        if not ok_reid:
            echecs.append("agent fake : réponse non ré-identifiée (prénom absent)")
        statut = "OK " if (ok_label and ok_reid) else "KO "
        print(f"{statut}[--] agent fake (pseudonymisation + ré-identification) "
              f"· outils={res['outils_utilises']}")

# --- Agent Phase 2 : action « préparer une relance » (fake, hors-ligne) ---
# Vérifie qu'une ACTION est proposée et RÉ-IDENTIFIÉE (vrai prénom, plus d'étiquette),
# sans aucun envoi réel (le fake ne fait qu'assembler l'objet action).
if employes:
    resA = agent_rh.run_agent([{"role": "user", "content": f"prépare une relance pour {prenom0}"}],
                              employes, A.executer_outil_agent, moteur="fake")
    acts = resA.get("actions", [])
    txt = " ".join(str(a) for a in acts)
    ok_act = bool(acts) and any(a.get("type") == "mailto" for a in acts)
    ok_reid = (prenom0 in txt) and ("Employé" not in txt)
    print(("OK " if (ok_act and ok_reid) else "KO ") + "[--] agent action relance (proposée + ré-identifiée)")
    if not (ok_act and ok_reid):
        echecs.append("agent action : relance non proposée/ré-identifiée")

# --- Phase 1 : crypto + extraction OCR + propositions (hors-ligne, données isolées) ---
# Chiffrement au repos : round-trip (sans clé -> clair ; avec clé -> enc:).
ok_crypto = crypto_rh.dechiffrer(crypto_rh.chiffrer("FR7630004")) == "FR7630004"
print(("OK " if ok_crypto else "KO ") + "[--] crypto_rh round-trip")
if not ok_crypto:
    echecs.append("crypto_rh : round-trip KO")

# Extraction : RIB -> IBAN, contrat -> dates, ARRÊT -> rien (aucune donnée de santé).
champs_rib = extraction_pj.extraire_champs("RIB / coordonnées bancaires",
                                           "IBAN FR76 3000 4000 0500 0012 3456 789 BIC X")
champs_contrat = extraction_pj.extraire_champs("Contrat de travail",
                                               "prend effet le 01/09/2026 jusqu'au 31/12/2026")
champs_arret = extraction_pj.extraire_champs("Arrêt de travail", "repos jusqu'au 20/06/2026 maladie")
ok_ext = (any(c["cible"] == "iban" for c in champs_rib)
          and any(c["cible"] == "profil:date_fin" for c in champs_contrat)
          and champs_arret == [])
print(("OK " if ok_ext else "KO ") + "[--] extraction (RIB+contrat extraits, arrêt ignoré)")
if not ok_ext:
    echecs.append("extraction_pj : champs inattendus")

# Propositions : ajout + appliquer, sur un fichier profils TEMPORAIRE (vraies données intactes).
if employes:
    email0 = employes[0]["email"]
    _orig_pf = A.PROFILS_FILE
    _tmp_pf = os.path.join(A.BASE_DIR, "_smoke_profils_tmp.json")
    A.PROFILS_FILE = _tmp_pf
    try:
        A.sauvegarder_profils({email0: {}})
        A._ajouter_propositions(email0, [
            {"cible": "profil:date_fin", "valeur": "31/12/2026",
             "apercu": "Fin : 31/12/2026", "libelle": "Fin de contrat", "chiffre": False},
            {"cible": "iban", "valeur": "…6789", "apercu": "IBAN", "libelle": "IBAN", "chiffre": True},
        ], "docSMOKE")
        prof = A.profil_de(email0)
        pid = next((p["id"] for p in prof.get("propositions", []) if p["cible"] == "profil:date_fin"), None)
        with client.session_transaction() as s:
            s["admin"] = True
            s["_csrf_token"] = "tok"
        client.post("/admin/employe/proposition/appliquer",
                    data={"email": email0, "prop_id": pid, "csrf_token": "tok"})
        prof2 = A.profil_de(email0)
        ok_prop = (len(prof.get("propositions", [])) == 2
                   and prof2.get("date_fin") == "31/12/2026"
                   and not any(p["id"] == pid for p in prof2.get("propositions", [])))
        print(("OK " if ok_prop else "KO ") + "[--] propositions (ajout + appliquer date_fin)")
        if not ok_prop:
            echecs.append("propositions : ajout/appliquer KO")
    finally:
        A.PROFILS_FILE = _orig_pf
        A._JSON_CACHE.pop(_tmp_pf, None)
        if os.path.exists(_tmp_pf):
            os.remove(_tmp_pf)

# --- Module Recrutement (hors-ligne + fiche candidat isolée) ---
import recrutement
import recrutement_ia
ct = extraction_pj.extraire_contact("Contact : 06 12 34 56 78 — a.b@exemple.fr")
an = recrutement_ia.analyser_cv("CV de test, préparateur 5 ans.", "Préparateur", moteur="fake")
ok_rec = (ct.get("email") == "a.b@exemple.fr" and ct.get("telephone")
          and isinstance(an.get("score"), int) and "resume" in an)
print(("OK " if ok_rec else "KO ") + "[--] recrutement (extraire_contact + analyse IA fake)")
if not ok_rec:
    echecs.append("recrutement : contact/analyse KO")

_orig_cf = recrutement.CANDIDATS_FILE
_tmp_cf = os.path.join(A.BASE_DIR, "_smoke_candidats_tmp.json")
recrutement.CANDIDATS_FILE = _tmp_cf
try:
    recrutement.sauvegarder_candidats({"cTEST": {
        "prenom": "Marie", "nom": "Leroy", "email": "marie.leroy@exemple.fr", "telephone": "",
        "poste_vise": "Préparateur", "source": "", "statut": "Reçu", "date_ajout": "01/01/2026",
        "notes_libres": "", "journal": [], "cv_texte": "Vaccination, LGPI.",
        "analyse_ia": {"score": 80, "adequation": "ok", "resume": "r", "points_forts": [], "points_attention": []}}})
    code = client.get("/admin/recrutement/candidat?id=cTEST").status_code
    print(("OK " if code == 200 else "KO ") + f"[{code}] /admin/recrutement/candidat (fiche)")
    if code != 200:
        echecs.append(f"fiche candidat (reçu {code})")
    # Agent recrutement (fake) : classement + action convocation (mailto vrai e-mail)
    import agent_recrutement
    r1 = agent_recrutement.run_agent_recrutement(
        [{"role": "user", "content": "classe les candidats"}],
        recrutement.executer_outil_recrutement, moteur="fake")
    r2 = agent_recrutement.run_agent_recrutement(
        [{"role": "user", "content": "prépare une convocation pour Marie"}],
        recrutement.executer_outil_recrutement, moteur="fake")
    ok_ag = ("classer_candidats" in r1.get("outils_utilises", [])
             and any(a.get("type") == "mailto" and a.get("to") == "marie.leroy@exemple.fr"
                     for a in r2.get("actions", [])))
    print(("OK " if ok_ag else "KO ") + "[--] agent recrutement (classement + convocation mailto)")
    if not ok_ag:
        echecs.append("agent recrutement : outil/action KO")
    # Lot C : grille d'entretien
    with client.session_transaction() as s:
        s["admin"] = True; s["_csrf_token"] = "tok"
    client.post("/admin/recrutement/grille",
                data={"id": "cTEST", "csrf_token": "tok", "crit_experience": "4", "crit_motivation": "5"})
    g = recrutement.charger_candidats()["cTEST"].get("grille", {})
    ok_g = g.get("experience") == 4 and g.get("motivation") == 5
    print(("OK " if ok_g else "KO ") + "[--] grille d'entretien (enregistrement)")
    if not ok_g:
        echecs.append("grille d'entretien KO")
finally:
    recrutement.CANDIDATS_FILE = _orig_cf
    A._JSON_CACHE.pop(_tmp_cf, None)
    if os.path.exists(_tmp_cf):
        os.remove(_tmp_cf)

# --- Lot A : embauche candidat -> salarié (TOTALEMENT isolé, vraies données intactes) ---
import tempfile
import csv as _csv
import shutil as _shutil
_td = tempfile.mkdtemp()
_sv = {"CF": recrutement.CANDIDATS_FILE, "CDD": recrutement.CANDIDATS_DOCS_DIR,
       "CDI": recrutement.CANDIDATS_DOCS_INDEX, "rDD": recrutement.DOCS_DIR,
       "EL": A.EMPLOYEES_LIVE, "EF": A.EMPLOYEES_FILE, "PF": A.PROFILS_FILE,
       "DD": A.DOCS_DIR, "DI": A.DOCS_INDEX}
try:
    recrutement.CANDIDATS_FILE = os.path.join(_td, "cand.json")
    recrutement.CANDIDATS_DOCS_DIR = os.path.join(_td, "cdocs"); os.makedirs(recrutement.CANDIDATS_DOCS_DIR)
    recrutement.CANDIDATS_DOCS_INDEX = os.path.join(recrutement.CANDIDATS_DOCS_DIR, "index.json")
    _docs = os.path.join(_td, "docs"); os.makedirs(_docs)
    recrutement.DOCS_DIR = _docs; A.DOCS_DIR = _docs; A.DOCS_INDEX = os.path.join(_docs, "index.json")
    A.EMPLOYEES_LIVE = os.path.join(_td, "emp_live.csv"); A.EMPLOYEES_FILE = A.EMPLOYEES_LIVE
    with open(A.EMPLOYEES_LIVE, "w", newline="", encoding="utf-8") as f:
        _csv.DictWriter(f, fieldnames=["nom", "prenom", "email"]).writeheader()
    A.PROFILS_FILE = os.path.join(_td, "profils.json")
    with open(os.path.join(recrutement.CANDIDATS_DOCS_DIR, "d1_cv.pdf"), "wb") as f:
        f.write(b"%PDF cv test")
    recrutement.sauvegarder_candidats_docs_index({"cE": [{"id": "d1", "fichier": "d1_cv.pdf",
        "nom_original": "cv.pdf", "type": "CV", "libelle": "CV", "taille": "1 Ko", "date_ajout": ""}]})
    recrutement.sauvegarder_candidats({"cE": {"prenom": "Eva", "nom": "Test", "email": "eva.test@ex.fr",
        "telephone": "0102030405", "poste_vise": "Préparateur", "source": "", "statut": "Retenu",
        "date_ajout": "", "notes_libres": "", "journal": [], "cv_texte": "", "analyse_ia": None}})
    with client.session_transaction() as s:
        s["admin"] = True; s["_csrf_token"] = "t"
    client.post("/admin/recrutement/embaucher", data={"id": "cE", "csrf_token": "t"})
    emp_ok = any(e["email"] == "eva.test@ex.fr" for e in A.charger_employes())
    prof_ok = A.profil_de("eva.test@ex.fr").get("poste") == "Préparateur"
    doc_ok = len(A.charger_docs_index().get("eva.test@ex.fr", [])) == 1
    zcode = client.get("/admin/recrutement/dossier-zip?id=cE").status_code
    ok_emb = emp_ok and prof_ok and doc_ok and zcode == 200
    print(("OK " if ok_emb else "KO ") + f"[--] embauche (salarié={emp_ok} profil={prof_ok} doc copié={doc_ok} zip={zcode})")
    if not ok_emb:
        echecs.append("embauche candidat->salarié KO")
finally:
    recrutement.CANDIDATS_FILE, recrutement.CANDIDATS_DOCS_DIR = _sv["CF"], _sv["CDD"]
    recrutement.CANDIDATS_DOCS_INDEX, recrutement.DOCS_DIR = _sv["CDI"], _sv["rDD"]
    A.EMPLOYEES_LIVE, A.EMPLOYEES_FILE, A.PROFILS_FILE = _sv["EL"], _sv["EF"], _sv["PF"]
    A.DOCS_DIR, A.DOCS_INDEX = _sv["DD"], _sv["DI"]
    for _p in [k for k in A._JSON_CACHE if _td.replace("\\", "/") in k.replace("\\", "/")]:
        A._JSON_CACHE.pop(_p, None)
    _shutil.rmtree(_td, ignore_errors=True)

# --- Lot B : /candidat_push (création candidat depuis le runner), isolé ---
_cf2, _cdd2, _cdi2 = recrutement.CANDIDATS_FILE, recrutement.CANDIDATS_DOCS_DIR, recrutement.CANDIDATS_DOCS_INDEX
_td2 = tempfile.mkdtemp()
recrutement.CANDIDATS_FILE = os.path.join(_td2, "c.json")
recrutement.CANDIDATS_DOCS_DIR = os.path.join(_td2, "cd")
recrutement.CANDIDATS_DOCS_INDEX = os.path.join(recrutement.CANDIDATS_DOCS_DIR, "index.json")
try:
    import base64 as _b64
    payload = {"prenom": "Léa", "nom": "Postulante", "email": "lea.p@ex.fr",
               "sujet": "Candidature préparateur", "filename": "cv.pdf",
               "content_b64": _b64.b64encode(b"%PDF cv test").decode(), "cv_texte": "CV de test"}
    st1 = json.loads(client.post(f"/candidat_push?cle={A.API_CLE}", json=payload).data).get("status")
    st2 = json.loads(client.post(f"/candidat_push?cle={A.API_CLE}", json=payload).data).get("status")
    r403 = client.post("/candidat_push?cle=mauvaise", json=payload).status_code
    nb = len(recrutement.charger_candidats())
    ok_b = st1 == "ajoute" and st2 == "doublon" and nb == 1 and r403 == 403
    print(("OK " if ok_b else "KO ") + f"[--] candidat_push (ajout={st1}, doublon={st2}, clé KO={r403})")
    if not ok_b:
        echecs.append("candidat_push KO")
finally:
    recrutement.CANDIDATS_FILE, recrutement.CANDIDATS_DOCS_DIR, recrutement.CANDIDATS_DOCS_INDEX = _cf2, _cdd2, _cdi2
    for _p in [k for k in A._JSON_CACHE if _td2.replace("\\", "/") in k.replace("\\", "/")]:
        A._JSON_CACHE.pop(_p, None)
    _shutil.rmtree(_td2, ignore_errors=True)

# --- Mail de bienvenue candidat (recrutement.mail_bienvenue, sans envoi) ---
_su_b, _cor_b = recrutement.mail_bienvenue(
    {"prenom": "Léa", "poste_vise": "Préparatrice en pharmacie", "email": "lea@ex.fr"})
_su_g, _cor_g = recrutement.mail_bienvenue({"prenom": "Léa", "poste_vise": "Rayonniste"})
ok_mb = ("Bienvenue" in _su_b and "Léa" in _cor_b
         and "diplôme de préparateur en pharmacie" in _cor_b
         and "dernier diplôme" in _cor_g
         and all(x in _cor_b for x in ("fiche signalétique", "carte Vitale", "RIB",
                                       "carte d'identité")))
from signature_mail import SIGNATURE
ok_mb = (ok_mb and SIGNATURE in _cor_b and SIGNATURE in _cor_g
         and "/docs-embauche/fiche-signaletique" in _cor_b
         and "/docs-embauche/mutuelle" in _cor_b and "mutuelle" in _cor_b)
print(("OK " if ok_mb else "KO ") + "[--] mail de bienvenue (préparateur + cas général)")
if not ok_mb:
    echecs.append("mail_bienvenue KO")

# --- Notifications congés : contenu des mails (conges_sender, sans SMTP) ---
import conges_sender
_base_cp = {"prenom": "Léa", "email": "lea@ex.fr", "debut": "03/08/2026",
            "fin": "08/08/2026", "nb": 6, "commentaire": "été", "motif_refus": ""}
_m_dep = conges_sender.construire_mail({**_base_cp, "action": "deposee"})
_m_acc = conges_sender.construire_mail({**_base_cp, "action": "acceptee"})
_m_ref = conges_sender.construire_mail({**_base_cp, "action": "refusee",
                                        "motif_refus": "effectifs insuffisants"})
ok_cs = (_m_dep is not None and _m_dep[0] == conges_sender.ADMIN_EMAIL
         and "Léa" in _m_dep[1] and "6 jours ouvrables" in _m_dep[2]
         and _m_acc is not None and _m_acc[0] == "lea@ex.fr" and "acceptés" in _m_acc[1]
         and _m_ref is not None and _m_ref[0] == "lea@ex.fr"
         and "effectifs insuffisants" in _m_ref[2]
         and conges_sender.construire_mail({"action": "annulee"}) is None
         and conges_sender.construire_mail({}) is None
         and SIGNATURE in _m_acc[2] and SIGNATURE in _m_ref[2])
print(("OK " if ok_cs else "KO ") + "[--] conges_sender.construire_mail (3 actions + vides)")
if not ok_cs:
    echecs.append("conges_sender.construire_mail KO")

# --- Dépôt d'une demande de congés (store isolé) : route OK + notification
#     best-effort silencieuse (pas de GITHUB_TOKEN → aucun appel réseau) ---
import planning_equipe as PE
_dcf = PE.DEMANDES_CP_FILE
_td3 = tempfile.mkdtemp()
PE.DEMANDES_CP_FILE = os.path.join(_td3, "demandes.json")
try:
    from datetime import date as _date, timedelta as _tdelta
    _d1 = (_date.today() + _tdelta(days=30)).isoformat()
    _d2 = (_date.today() + _tdelta(days=35)).isoformat()
    _rc = client.post("/mon-espace/conges/demander",
                      data={"token": tok0, "debut": _d1, "fin": _d2,
                            "commentaire": "smoke"}).status_code
    _nbd = len(PE.charger_demandes_cp())
    ok_dm = _rc == 302 and _nbd == 1
    print(("OK " if ok_dm else "KO ") + f"[--] dépôt demande congés (code={_rc}, enregistrées={_nbd})")
    if not ok_dm:
        echecs.append("dépôt demande congés KO")
finally:
    PE.DEMANDES_CP_FILE = _dcf
    for _p in [k for k in A._JSON_CACHE if _td3.replace("\\", "/") in k.replace("\\", "/")]:
        A._JSON_CACHE.pop(_p, None)
    _shutil.rmtree(_td3, ignore_errors=True)

# --- Paie : ventilation par semaine + sujétion (garde dim./férié) + mail
#     comptable HTML (comptable_sender, sans SMTP ni réseau) ---
import comptable_sender as CS
_mj = A.calculer_majorations([{"label": "Mar 07/07", "plus": 10, "moins": 0},
                              {"label": "Dim 12/07", "plus": 10, "moins": 0}], 35.0, 7, 2026)
_mf = A.calculer_majorations([{"label": "Mar 14/07", "plus": 7, "moins": 7}], 35.0, 7, 2026)
_res = [{"prenom": "Test", "nom": "Un", "contrat_hebdo": 35.0, "statut": "ok",
         "plus": 20.0, "moins": 0.0, "solde": 20.0, "valide": True,
         "saisi_par_admin": False, "corrige": False, "commentaire": "", **_mj},
        {"prenom": "Test", "nom": "Deux", "contrat_hebdo": 35.0, "statut": "manquant"}]
_sj, _txt, _html = CS.construire_mail_comptable(_res, "Juillet 2026")
ok_pv = (_mj["sup25"] == 8.0 and _mj["sup50"] == 12.0 and _mj["sujetion"] == 10.0
         and _mj["semaines"][0]["lundi"] == "2026-07-06"
         and _mj["semaines"][0]["plus"] == 20.0
         and _mj["semaines"][0]["sujetion_jours"] == ["dim. 12/07"]
         and _mf["sup25"] == 0 and _mf["sujetion"] == 7.0
         and "férié 14/07 (Fête nationale)" in _mf["semaines"][0]["sujetion_jours"]
         and "Juillet 2026" in _sj and "sujétion 10h" in _txt
         and "du lun 06/07 au dim 12/07" in _html and "dim. 12/07" in _html
         and "Total équipe" in _html and "Test DEUX" in _html and SIGNATURE in _txt)
print(("OK " if ok_pv else "KO ") + "[--] paie : semaines + sujétion + mail comptable HTML")
if not ok_pv:
    echecs.append("paie semaines/sujétion/mail comptable KO")

# --- Aperçu comptable : la page répond (200) ou renvoie vers /admin/mois avec
#     un motif (302 — relevés non validés/vides selon les données du poste) ;
#     ajustements : champ modifié appliqué + totaux recalculés + flag ajuste,
#     champ vide/illisible ignoré, statut sans_contrat ventilé à la main ---
_ca = client.get("/admin/comptable/apercu")
ok_ap = _ca.status_code in (200, 302)
_it1 = {"prenom": "Test", "nom": "Un", "email": "t1@ex.fr", "contrat_hebdo": 35.0,
        "statut": "ok", "plus": 20.0, "moins": 0.0, "solde": 20.0,
        **A.calculer_majorations([{"label": "Mar 07/07", "plus": 10, "moins": 0},
                                  {"label": "Dim 12/07", "plus": 10, "moins": 0}], 35.0, 7, 2026)}
_it2 = {"prenom": "Test", "nom": "Trois", "email": "t3@ex.fr", "contrat_hebdo": 0,
        "statut": "sans_contrat", "plus": 4.0, "moins": 0.0, "solde": 4.0}
_resume2 = [_it1, _it2]
A._appliquer_ajustements_comptable(_resume2, {
    "v_t1@ex.fr_2026-07-06_sujetion": "8",      # correction d'une garde surévaluée
    "v_t1@ex.fr_2026-07-06_plus": "",            # vide -> conservé
    "v_t1@ex.fr_2026-07-06_sup50": "abc",        # illisible -> conservé
    "v_t3@ex.fr_tot_sup25": "4",                 # ventilation manuelle
})
ok_aj = (_it1["ajuste"] and _it1["semaines"][0]["sujetion"] == 8.0
         and _it1["sujetion"] == 8.0 and _it1["plus"] == 20.0
         and _it1["semaines"][0]["sup50"] == 12.0
         and _it2["ajuste"] and _it2["statut"] == "ok" and _it2["sup25"] == 4.0)
_sj2, _txt2, _html2 = CS.construire_mail_comptable(_resume2, "Juillet 2026")
ok_aj = ok_aj and "chiffres ajustés par la pharmacie" in _html2
print(("OK " if ok_ap and ok_aj else "KO ")
      + f"[{_ca.status_code}] aperçu comptable + ajustements (flag ajuste dans le mail)")
if not (ok_ap and ok_aj):
    echecs.append("aperçu/ajustements comptable KO")

# --- Congés payés pris sur la période de paie (stores planning isolés) :
#     plage 13-25/07 = 11 j ouvrables (dim 19 + férié 14 exclus), ponctuel CP
#     06/07 = +1 j, ponctuel un férié (14/07) ignoré ; affichage mail (section
#     collaborateur + relevé manquant en congés) ---
_tc = tempfile.mkdtemp()
_absf0, _chgf0 = PE.ABSENCES_FILE, PE.CHANGEMENTS_FILE
PE.ABSENCES_FILE = os.path.join(_tc, "abs.json")
PE.CHANGEMENTS_FILE = os.path.join(_tc, "chg.json")
try:
    with open(PE.ABSENCES_FILE, "w", encoding="utf-8") as f:
        json.dump([{"id": 1, "email": "t1@ex.fr", "debut": "2026-07-13",
                    "fin": "2026-07-25", "motif": "Congés payés"}], f)
    with open(PE.CHANGEMENTS_FILE, "w", encoding="utf-8") as f:
        json.dump({"2026-07-06": {"t1@ex.fr": {"motif": "Congés payés", "creneaux": []}},
                   "2026-07-14": {"t1@ex.fr": {"motif": "Congés payés", "creneaux": []}}}, f)
    _cg = A._conges_paie(7, 2026)
finally:
    PE.ABSENCES_FILE, PE.CHANGEMENTS_FILE = _absf0, _chgf0
    for _p in [k for k in A._JSON_CACHE if _tc.replace("\\", "/") in k.replace("\\", "/")]:
        A._JSON_CACHE.pop(_p, None)
    _shutil.rmtree(_tc, ignore_errors=True)
_c1 = _cg.get("t1@ex.fr") or {"plages": [], "total": 0}
ok_cg = (_c1["total"] == 12 and len(_c1["plages"]) == 2
         and _c1["plages"][0] == {"debut": "06/07", "fin": "06/07", "jours": 1}
         and _c1["plages"][1] == {"debut": "13/07", "fin": "25/07", "jours": 11})
_it_abs = {"prenom": "Nora", "nom": "L", "email": "n@ex.fr", "contrat_hebdo": 35.0,
           "statut": "manquant",
           "conges": {"plages": [{"debut": "24/06", "fin": "25/07", "jours": 27}], "total": 27}}
_sj3, _txt3, _html3 = CS.construire_mail_comptable([{**_it1, "conges": _c1}, _it_abs],
                                                   "Juillet 2026")
ok_cg = (ok_cg and "Congés payés pris sur la période" in _html3
         and "du 13/07 au 25/07 (11 j)" in _html3
         and "le 06/07 (1 j) · du 13/07 au 25/07 (11 j) — total 12 j ouvrables" in _txt3
         and "Nora L (congés payés : du 24/06 au 25/07 (27 j))" in _html3)
print(("OK " if ok_cg else "KO ") + "[--] congés payés période de paie (calcul + mail)")
if not ok_cg:
    echecs.append("congés payés période de paie KO")

if cree_temp and os.path.exists(fichier_temp):
    os.remove(fichier_temp)

if echecs:
    print(f"\n[ECHEC] {len(echecs)} route(s) en erreur :")
    for x in echecs:
        print("   -", x)
    sys.exit(1)
print("\n[OK] Toutes les pages repondent correctement.")
