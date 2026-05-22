import os
import glob
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from parse_planning import parse_planning
from parse_releve import parse_releve

MOIS_FR = {
    1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril",
    5: "Mai", 6: "Juin", 7: "Juillet", 8: "Août",
    9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Décembre"
}

PLANNING_FOLDER = "planning"
REPONSES_FOLDER = "reponses"
RESUMES_FOLDER  = "resumes"


def trouver_planning():
    fichiers = glob.glob(os.path.join(PLANNING_FOLDER, "*.html"))
    if not fichiers:
        raise FileNotFoundError("Aucun fichier HTML dans le dossier planning/")
    return fichiers[0]


def trouver_releves():
    releves = {}
    for ext in ("*.docx", "*.doc", "*.pdf"):
        for path in glob.glob(os.path.join(REPONSES_FOLDER, ext)):
            nom_fichier = os.path.basename(path)
            prenom = nom_fichier.split("_")[0].capitalize()
            releves[prenom] = path
    return releves


def generer_comparaison():
    os.makedirs(RESUMES_FOLDER, exist_ok=True)
    mois = datetime.now().month
    annee = datetime.now().year
    mois_annee = f"{MOIS_FR[mois]} {annee}"

    planning = parse_planning(trouver_planning())
    releves  = trouver_releves()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Comparaison {mois_annee}"

    # Titre
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Comparaison heures — {mois_annee}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    # En-têtes
    headers = ["Prénom", "Planning\n(total)", "Déclaré\n(h+)", "Déclaré\n(h−)", "Solde\ndéclaré", "Écart\navec planning", "Statut", "Relevé reçu ?"]
    ws.append([])
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="2E75B6")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[3].height = 30

    col_widths = [14, 14, 12, 12, 12, 16, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Données
    for emp in planning:
        prenom = emp["prenom"]
        total_planning = emp["total"]

        releve_recu = prenom in releves
        if releve_recu:
            r = parse_releve(releves[prenom])
            h_plus   = r["heures_plus"]
            h_moins  = r["heures_moins"]
            solde    = r["solde"]
            ecart    = round(solde - total_planning, 2)
            if abs(ecart) <= 0.5:
                statut = "OK"
                statut_fill = PatternFill("solid", fgColor="C6EFCE")
                statut_color = "276221"
            else:
                statut = "À VÉRIFIER"
                statut_fill = PatternFill("solid", fgColor="FFCCCC")
                statut_color = "9C0006"
        else:
            h_plus = h_moins = solde = ecart = ""
            statut = "EN ATTENTE"
            statut_fill = PatternFill("solid", fgColor="FFEB9C")
            statut_color = "9C6500"

        row = [
            prenom,
            total_planning,
            h_plus,
            h_moins,
            solde,
            ecart,
            statut,
            "Oui" if releve_recu else "Non"
        ]
        ws.append(row)
        r_idx = ws.max_row

        for col, cell in enumerate(ws[r_idx], 1):
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if col == 7:
                cell.fill = statut_fill
                cell.font = Font(bold=True, color=statut_color)
            elif col == 6 and ecart != "":
                cell.font = Font(color="9C0006" if ecart < -0.5 or ecart > 0.5 else "276221")

    path = os.path.join(RESUMES_FOLDER, f"Comparaison_{MOIS_FR[mois]}_{annee}.xlsx")
    wb.save(path)
    print(f"Rapport généré : {path}")
    return path


if __name__ == "__main__":
    generer_comparaison()
