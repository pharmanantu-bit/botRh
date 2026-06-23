import json
import os
import csv
import io
import uuid
import hashlib
from datetime import datetime
from flask import Flask, request, render_template, abort, redirect, url_for, session, send_file
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Dossier du projet — sert d'ancrage pour tous les chemins de fichiers, car
# sous le serveur WSGI (PythonAnywhere) le répertoire courant n'est pas celui
# du projet : sans ça, employees.csv, documents/, reponses_*.json sont introuvables.
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Configuration sensible chargée depuis le .env (avec valeurs par défaut pour
# ne rien casser). Pour durcir la sécurité, définir ces clés dans le .env du
# serveur — en priorité ADMIN_PASSWORD avec un mot de passe fort.
from dotenv import load_dotenv
load_dotenv(os.path.join(BASE_DIR, ".env"))

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "botRh-admin-2026")
# Limite de taille des fichiers uploadés (documents RH) : 12 Mo.
app.config["MAX_CONTENT_LENGTH"] = 12 * 1024 * 1024

# Journalisation des erreurs dans logs/erreurs.log (Flask y écrit aussi
# automatiquement les exceptions non gérées) — consultable via /admin/erreurs.
import logging
os.makedirs(os.path.join(BASE_DIR, "logs"), exist_ok=True)
_err_handler = logging.FileHandler(os.path.join(BASE_DIR, "logs", "erreurs.log"), encoding="utf-8")
_err_handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s: %(message)s"))
app.logger.addHandler(_err_handler)
app.logger.setLevel(logging.INFO)

SECRET = os.getenv("TOKEN_SECRET", "pharmacie-nanterre-2026")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "pharma92")
API_CLE = os.getenv("API_CLE", "botRh-trigger-2026")
# Jour de clôture RÉELLE de la saisie (côté interne) : au-delà, le formulaire
# est fermé. On communique le 25 aux employés (mails + formulaire) mais on
# accepte en réalité jusqu'au 28, pour garder une marge. Surchargeable par env.
JOUR_CLOTURE = int(os.getenv("JOUR_CLOTURE", "28"))
# employees.csv (versionné) = liste de départ ; employees_live.csv (gitignore)
# = liste gérée par l'admin sur le serveur. On lit le live s'il existe, et c'est
# lui qui fait foi (évite tout conflit de déploiement avec git).
EMPLOYEES_FILE = os.path.join(BASE_DIR, "employees.csv")
EMPLOYEES_LIVE = os.path.join(BASE_DIR, "employees_live.csv")

def employees_path(pour_ecriture=False):
    if pour_ecriture:
        return EMPLOYEES_LIVE
    return EMPLOYEES_LIVE if os.path.exists(EMPLOYEES_LIVE) else EMPLOYEES_FILE

def reponses_file(mois=None, annee=None):
    if mois is None:
        mois = datetime.now().month
    if annee is None:
        annee = datetime.now().year
    return os.path.join(BASE_DIR, f"reponses_{mois}_{annee}.json")
UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "planning_img")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

MOIS_FR = {
    1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril",
    5: "Mai", 6: "Juin", 7: "Juillet", 8: "Août",
    9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Décembre"
}


from tokens import generer_token, resoudre_employe, reponse_de, tokens_valides


def charger_reponses(mois=None, annee=None):
    f = reponses_file(mois, annee)
    if os.path.exists(f):
        with open(f, encoding="utf-8") as fp:
            return json.load(fp)
    return {}


def sauvegarder_reponse(token, data):
    f = reponses_file()
    reponses = charger_reponses()
    reponses[token] = data
    with open(f, "w", encoding="utf-8") as fp:
        json.dump(reponses, fp, ensure_ascii=False, indent=2)


def ecrire_reponses(reponses, mois=None, annee=None):
    with open(reponses_file(mois, annee), "w", encoding="utf-8") as fp:
        json.dump(reponses, fp, ensure_ascii=False, indent=2)


@app.route("/releve")
def formulaire():
    token = request.args.get("token", "")
    prenom = request.args.get("prenom", "")
    if not token or not prenom:
        abort(404)

    import calendar
    mois = datetime.now().month
    annee = datetime.now().year
    mois_annee = f"{MOIS_FR[mois]} {annee}"

    mois_prec = 12 if mois == 1 else mois - 1
    annee_prec = annee - 1 if mois == 1 else annee
    nb_jours_prec = calendar.monthrange(annee_prec, mois_prec)[1]
    nb_jours_mois = calendar.monthrange(annee, mois)[1]

    jours_prec = list(range(24, nb_jours_prec + 1))
    jours_mois = list(range(1, nb_jours_mois + 1))

    JOURS_FR = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]
    from datetime import date as dt_date

    noms_jours_prec = {j: JOURS_FR[dt_date(annee_prec, mois_prec, j).weekday()] for j in jours_prec}
    noms_jours_mois = {j: JOURS_FR[dt_date(annee, mois, j).weekday()] for j in jours_mois}
    weekend_prec = {j for j in jours_prec if dt_date(annee_prec, mois_prec, j).weekday() >= 5}
    weekend_mois = {j for j in jours_mois if dt_date(annee, mois, j).weekday() >= 5}

    reponses = charger_reponses()
    emp = resoudre_employe(token, charger_employes())
    token_canon = generer_token(emp["prenom"], emp["email"]) if emp else token
    deja_rempli = (reponse_de(reponses, emp["prenom"], emp["email"]) is not None) if emp else (token_canon in reponses)
    modifiable = datetime.now().day <= JOUR_CLOTURE

    return render_template("form.html",
        prenom=prenom,
        token=token,
        mois_annee=mois_annee,
        mois_nom=MOIS_FR[mois].upper(),
        mois_prec_nom=MOIS_FR[mois_prec].upper(),
        jours_prec=jours_prec,
        jours_mois=jours_mois,
        noms_jours_prec=noms_jours_prec,
        noms_jours_mois=noms_jours_mois,
        weekend_prec=weekend_prec,
        weekend_mois=weekend_mois,
        deja_rempli=deja_rempli,
        modifiable=modifiable,
        jour_cloture=JOUR_CLOTURE
    )


def extraire_detail_jours(form, mois, annee):
    """Reconstruit le détail jour par jour saisi dans le formulaire de relevé.
    Renvoie une liste {label, plus, moins} pour les seuls jours ayant des heures.
    La période couvre du 24 du mois précédent à la fin du mois courant
    (mêmes champs que form.html : prec_plus_J / prec_moins_J / mois_plus_J / mois_moins_J)."""
    import calendar
    from datetime import date as dt_date
    JOURS_FR = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]
    mois_prec = 12 if mois == 1 else mois - 1
    annee_prec = annee - 1 if mois == 1 else annee
    nb_jours_prec = calendar.monthrange(annee_prec, mois_prec)[1]
    nb_jours_mois = calendar.monthrange(annee, mois)[1]

    def lire(champ):
        try:
            return float(form.get(champ, 0) or 0)
        except ValueError:
            return 0.0

    jours = []
    for j in range(24, nb_jours_prec + 1):
        p, m = lire(f"prec_plus_{j}"), lire(f"prec_moins_{j}")
        if p or m:
            nom = JOURS_FR[dt_date(annee_prec, mois_prec, j).weekday()]
            jours.append({"label": f"{nom} {j:02d}/{mois_prec:02d}", "plus": p, "moins": m})
    for j in range(1, nb_jours_mois + 1):
        p, m = lire(f"mois_plus_{j}"), lire(f"mois_moins_{j}")
        if p or m:
            nom = JOURS_FR[dt_date(annee, mois, j).weekday()]
            jours.append({"label": f"{nom} {j:02d}/{mois:02d}", "plus": p, "moins": m})
    return jours


@app.route("/envoyer", methods=["POST"])
def envoyer():
    token = request.form.get("token", "")
    prenom = request.form.get("prenom", "")
    heures_plus = request.form.get("heures_plus", "0")
    heures_moins = request.form.get("heures_moins", "0")
    commentaire = request.form.get("commentaire", "")
    date_signature = request.form.get("date_signature", "")
    signature = request.form.get("signature", "")

    if not token or not prenom or not date_signature or not signature:
        abort(400)

    # Valider le jeton et identifier l'employé (bloque les soumissions forgées),
    # puis stocker sous le jeton canonique (nouveau) pour des recherches fiables.
    emp = resoudre_employe(token, charger_employes())
    if not emp:
        abort(403)
    # Blocage réel au JOUR_CLOTURE (28, côté interne) mais on communique le 25
    # aux employés : ce message ne s'affiche qu'au-delà du 28 (saisie vraiment close).
    if datetime.now().day > JOUR_CLOTURE:
        return ("La période de saisie de ce mois est clôturée (date limite : le 25). "
                "Vos heures effectuées seront comptabilisées le mois suivant."), 403
    prenom = emp["prenom"]
    token = generer_token(prenom, emp["email"])

    mois = datetime.now().month
    annee = datetime.now().year
    jours_detail = extraire_detail_jours(request.form, mois, annee)

    sauvegarder_reponse(token, {
        "prenom": prenom,
        "heures_plus": float(heures_plus or 0),
        "heures_moins": float(heures_moins or 0),
        "commentaire": commentaire,
        "date_signature": date_signature,
        "signature": signature,
        "date": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "mois": mois,
        "annee": annee,
        "jours": jours_detail,
    })

    try:
        # NB : repository_dispatch limite client_payload à 10 propriétés max.
        # On regroupe donc mois+annee dans "periode" pour rester sous la limite.
        notifier_releve({
            "prenom": prenom,
            "email": emp["email"],
            "heures_plus": heures_plus,
            "heures_moins": heures_moins,
            "commentaire": commentaire,
            "date_signature": date_signature,
            "signature": signature,
            "date": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "periode": [mois, annee],
            "jours": jours_detail,
        })
    except Exception:
        app.logger.exception("Échec du déclenchement de la notification (notifier_releve)")

    return render_template("merci.html", prenom=prenom,
        heures_plus=heures_plus, heures_moins=heures_moins,
        mois_annee=f"{MOIS_FR[mois]} {annee}"
    )


def planning_file(mois=None, annee=None):
    if mois is None:
        mois = datetime.now().month
    if annee is None:
        annee = datetime.now().year
    return os.path.join(BASE_DIR, f"planning_{mois}_{annee}.json")

def charger_planning(mois=None, annee=None):
    f = planning_file(mois, annee)
    if os.path.exists(f):
        with open(f, encoding="utf-8") as fp:
            return json.load(fp)
    return {}

def sauvegarder_planning(data):
    with open(planning_file(), "w", encoding="utf-8") as fp:
        json.dump(data, fp, ensure_ascii=False, indent=2)

def hm_to_float(s):
    s = s.strip().replace("min", "").replace(" ", "")
    if "h" in s:
        parts = s.split("h")
        h = float(parts[0] or 0)
        m = float(parts[1] or 0) if len(parts) > 1 else 0
        return round(h + m / 60, 2)
    return 0.0

def charger_employes():
    employes = []
    chemin = employees_path()
    if os.path.exists(chemin):
        with open(chemin, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                employes.append({"prenom": row["prenom"], "nom": row["nom"], "email": row["email"]})
    return employes


@app.route("/admin", methods=["GET", "POST"])
def admin():
    if request.method == "POST":
        mdp = request.form.get("password", "")
        if mdp == ADMIN_PASSWORD:
            session["admin"] = True
        else:
            return render_template("admin_login.html", erreur=True)

    if not session.get("admin"):
        return render_template("admin_login.html", erreur=False)

    mois_courant, annee_courante = datetime.now().month, datetime.now().year
    mois = int(request.args.get("mois", mois_courant))
    annee = int(request.args.get("annee", annee_courante))
    est_courant = (mois == mois_courant and annee == annee_courante)

    reponses = charger_reponses(mois, annee)
    employes = charger_employes()
    planning = charger_planning(mois, annee)

    # Navigation mois précédent / suivant (sans dépasser le mois courant)
    mois_prec = 12 if mois == 1 else mois - 1
    annee_prec = annee - 1 if mois == 1 else annee
    mois_suiv = 1 if mois == 12 else mois + 1
    annee_suiv = annee + 1 if mois == 12 else annee
    peut_suivant = not (annee > annee_courante or (annee == annee_courante and mois >= mois_courant))

    # Image de planning du mois
    img_nom = f"planning_{mois}_{annee}.png"
    img_url = f"/static/planning_img/{img_nom}" if os.path.exists(os.path.join(UPLOAD_FOLDER, img_nom)) else None

    resultats = []
    for emp in employes:
        token = generer_token(emp["prenom"], emp["email"])
        reponse = reponse_de(reponses, emp["prenom"], emp["email"])
        plan = planning.get(emp["prenom"], {})
        plan_total = plan.get("total", None)

        if reponse:
            solde = round(reponse["heures_plus"] - reponse["heures_moins"], 2)
            if plan_total is not None:
                ecart = round(solde - plan_total, 2)
                if abs(ecart) <= 0.5:
                    statut_comp = "OK"
                else:
                    statut_comp = "A VERIFIER"
            else:
                ecart = None
                statut_comp = None
        else:
            solde = None
            ecart = None
            statut_comp = None

        resultats.append({
            "prenom": emp["prenom"],
            "nom": emp["nom"],
            "email": emp["email"],
            "repondu": reponse is not None,
            "valide": reponse.get("valide", False) if reponse else False,
            "heures_plus": reponse["heures_plus"] if reponse else "-",
            "heures_moins": reponse["heures_moins"] if reponse else "-",
            "solde": solde if solde is not None else "-",
            "commentaire": reponse.get("commentaire", "") if reponse else "",
            "date": reponse["date"] if reponse else "-",
            "plan_trame": plan.get("trame", "-"),
            "plan_total": plan_total if plan_total is not None else "-",
            "plan_absences": plan.get("absences", ""),
            "plan_jours": plan.get("jours", "-"),
            "ecart": ecart if ecart is not None else "-",
            "statut_comp": statut_comp,
            "lien": f"/releve?token={token}&prenom={emp['prenom']}",
        })

    # Détail complet par employé (clé = email) pour la fenêtre de détail.
    # "jours" est vide pour les relevés antérieurs à la sauvegarde du détail.
    releves_detail = {}
    for emp in employes:
        reponse = reponse_de(reponses, emp["prenom"], emp["email"])
        if reponse:
            releves_detail[emp["email"]] = {
                "commentaire": reponse.get("commentaire", ""),
                "signature": reponse.get("signature", ""),
                "date_signature": reponse.get("date_signature", ""),
                "date": reponse.get("date", ""),
                "jours": reponse.get("jours", []),
            }

    repondus = sum(1 for r in resultats if r["repondu"])
    return render_template("admin.html", resultats=resultats, repondus=repondus,
                           total=len(resultats), mois=mois, annee=annee,
                           mois_annee=f"{MOIS_FR[mois]} {annee}",
                           est_courant=est_courant, planning=planning, img_url=img_url,
                           a_planning=(planning != {}),
                           nav_prec={"mois": mois_prec, "annee": annee_prec},
                           nav_suiv=({"mois": mois_suiv, "annee": annee_suiv} if peut_suivant else None),
                           releves_detail=releves_detail)


def sauvegarder_employes(employes):
    with open(employees_path(pour_ecriture=True), "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["nom", "prenom", "email"])
        writer.writeheader()
        writer.writerows(employes)


# --- Dossier salarié (données personnelles, fichier gitignoré, source = serveur) ---
PROFILS_FILE = os.path.join(BASE_DIR, "profils_rh.json")

# Champs du profil RH (clé interne -> libellé affiché)
CHAMPS_PROFIL = [
    ("poste", "Poste / fonction"),
    ("type_contrat", "Type de contrat"),
    ("date_entree", "Date d'entrée"),
    ("date_fin", "Fin de contrat (si CDD)"),
    ("fin_essai", "Fin de période d'essai"),
    ("visite_medicale", "Prochaine visite médicale"),
    ("telephone", "Téléphone"),
    ("naissance", "Date de naissance"),
    ("adresse", "Adresse"),
    ("urgence_nom", "Contact d'urgence (nom)"),
    ("urgence_tel", "Contact d'urgence (tél.)"),
    ("notes", "Notes RH (interne)"),
]

# Documents obligatoires pour un dossier complet (checklist)
DOCS_REQUIS = ["Contrat de travail", "RIB / coordonnées bancaires", "Pièce d'identité"]

def parse_date_fr(s):
    """Parse 'jj/mm/aaaa' (ou jj-mm-aaaa) -> date, sinon None."""
    if not s:
        return None
    s = s.strip().replace("-", "/").replace(".", "/")
    from datetime import datetime as _dt
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return _dt.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def alertes_employe(profil):
    """Liste d'alertes/échéances d'un employé, calculées depuis son profil.
    Chaque alerte : {niveau: 'rouge'|'orange'|'info', texte: ...}."""
    if not profil:
        return []
    auj = datetime.now().date()
    alertes = []

    # Fin de CDD
    fin = parse_date_fr(profil.get("date_fin"))
    if fin and (profil.get("type_contrat") in ("CDD", "Intérim", "Apprentissage", "Stage") or profil.get("date_fin")):
        j = (fin - auj).days
        if j < 0:
            alertes.append({"niveau": "rouge", "texte": f"Contrat terminé le {fin:%d/%m/%Y}"})
        elif j <= 30:
            alertes.append({"niveau": "orange", "texte": f"Fin de contrat dans {j} j ({fin:%d/%m/%Y})"})

    # Fin de période d'essai
    essai = parse_date_fr(profil.get("fin_essai"))
    if essai:
        j = (essai - auj).days
        if 0 <= j <= 15:
            alertes.append({"niveau": "orange", "texte": f"Fin de période d'essai dans {j} j ({essai:%d/%m/%Y})"})

    # Visite médicale
    vm = parse_date_fr(profil.get("visite_medicale"))
    if vm:
        j = (vm - auj).days
        if j < 0:
            alertes.append({"niveau": "rouge", "texte": f"Visite médicale dépassée ({vm:%d/%m/%Y})"})
        elif j <= 30:
            alertes.append({"niveau": "orange", "texte": f"Visite médicale dans {j} j ({vm:%d/%m/%Y})"})

    # Anniversaire d'ancienneté
    entree = parse_date_fr(profil.get("date_entree"))
    if entree:
        try:
            prochain = entree.replace(year=auj.year)
        except ValueError:
            prochain = entree.replace(year=auj.year, day=28)
        if prochain < auj:
            try:
                prochain = entree.replace(year=auj.year + 1)
            except ValueError:
                prochain = entree.replace(year=auj.year + 1, day=28)
        j = (prochain - auj).days
        annees = prochain.year - entree.year
        if 0 <= j <= 30 and annees > 0:
            alertes.append({"niveau": "info", "texte": f"{annees} an(s) d'ancienneté le {prochain:%d/%m/%Y}"})

    return alertes

def docs_manquants(email):
    """Documents requis encore absents pour cet employé."""
    types_presents = {d.get("type") for d in docs_de(email)}
    return [t for t in DOCS_REQUIS if t not in types_presents]

# Types d'événements du journal RH
TYPES_EVENEMENT = ["Entretien", "Augmentation", "Avertissement",
                   "Changement de poste", "Formation", "Congés", "Autre"]

# Checklists d'arrivée (onboarding) et de départ (offboarding)
TACHES_ARRIVEE = ["Contrat signé", "RIB reçu", "Badge / clés remis", "Blouse fournie",
                  "Accès logiciel créé", "Visite médicale d'embauche", "Présentation à l'équipe"]
TACHES_DEPART = ["Badge / clés récupérés", "Blouse rendue", "Accès logiciel désactivés",
                 "Solde de tout compte", "Attestation employeur remise", "Certificat de travail remis"]

# Dossier des photos de profil (gitignoré, données perso)
PHOTOS_DIR = os.path.join(BASE_DIR, "photos_rh")

def alertes_documents(email):
    """Alertes sur les documents qui expirent (champ 'expiration')."""
    auj = datetime.now().date()
    out = []
    for d in docs_de(email):
        exp = parse_date_fr(d.get("expiration"))
        if not exp:
            continue
        j = (exp - auj).days
        libelle = d.get("libelle") or d.get("type") or "Document"
        if j < 0:
            out.append({"niveau": "rouge", "texte": f"{libelle} expiré ({exp:%d/%m/%Y})"})
        elif j <= 45:
            out.append({"niveau": "orange", "texte": f"{libelle} expire dans {j} j ({exp:%d/%m/%Y})"})
    return out

def anniversaire_proche(profil):
    """Alerte anniversaire (date de naissance) dans les 14 jours."""
    naiss = parse_date_fr(profil.get("naissance")) if profil else None
    if not naiss:
        return []
    auj = datetime.now().date()
    try:
        prochain = naiss.replace(year=auj.year)
    except ValueError:
        prochain = naiss.replace(year=auj.year, day=28)
    if prochain < auj:
        try:
            prochain = naiss.replace(year=auj.year + 1)
        except ValueError:
            prochain = naiss.replace(year=auj.year + 1, day=28)
    j = (prochain - auj).days
    if 0 <= j <= 14:
        age = prochain.year - naiss.year
        quand = "aujourd'hui 🎂" if j == 0 else f"dans {j} j"
        return [{"niveau": "info", "texte": f"Anniversaire {quand} ({age} ans, le {prochain:%d/%m})"}]
    return []

def alertes_completes(email, profil):
    """Toutes les alertes d'un employé : échéances profil + documents + anniversaire."""
    return alertes_employe(profil) + alertes_documents(email) + anniversaire_proche(profil)

def journal_trie(journal):
    """Trie les événements du journal du plus récent au plus ancien."""
    from datetime import date as _d
    return sorted(journal, key=lambda e: parse_date_fr(e.get("date")) or _d(1900, 1, 1), reverse=True)

def charger_profils():
    if os.path.exists(PROFILS_FILE):
        with open(PROFILS_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}

def sauvegarder_profils(profils):
    with open(PROFILS_FILE, "w", encoding="utf-8") as f:
        json.dump(profils, f, ensure_ascii=False, indent=2)

def profil_de(email):
    return charger_profils().get(email, {})


# --- Documents RH (privés admin, stockés sur disque, gitignorés) ---
DOCS_DIR = os.path.join(BASE_DIR, "documents_rh")
DOCS_INDEX = os.path.join(DOCS_DIR, "index.json")
EXT_DOCS_OK = {".pdf", ".png", ".jpg", ".jpeg", ".docx", ".doc"}

# Familles de documents RH -> sous-types. Sert au menu déroulant (optgroups)
# et au classement des documents par famille dans la fiche.
FAMILLES_DOCS = {
    "Contractuel & paie": [
        "Contrat de travail", "Avenant", "Bulletin de paie",
        "Solde de tout compte", "RIB / coordonnées bancaires",
    ],
    "Identité & administratif": [
        "Pièce d'identité", "Titre de séjour", "Carte vitale",
        "Justificatif de domicile", "N° de Sécurité sociale",
    ],
    "Santé & formation": [
        "Visite médicale", "Arrêt de travail", "Diplôme",
        "Certification / habilitation", "Attestation de formation",
    ],
    "Suivi RH & divers": [
        "Entretien annuel / professionnel", "Attestation employeur",
        "Courrier", "Avertissement", "Autre / divers",
    ],
}

def famille_du_type(type_doc):
    """Retrouve la famille d'un sous-type ; 'Suivi RH & divers' par défaut/legacy."""
    for fam, sous in FAMILLES_DOCS.items():
        if type_doc in sous:
            return fam
    return "Suivi RH & divers"

def grouper_docs_par_famille(docs):
    """Regroupe les documents d'un employé par famille, dans l'ordre des familles."""
    groupes = {fam: [] for fam in FAMILLES_DOCS}
    for d in docs:
        groupes[famille_du_type(d.get("type", ""))].append(d)
    return {fam: items for fam, items in groupes.items() if items}

def charger_docs_index():
    if os.path.exists(DOCS_INDEX):
        with open(DOCS_INDEX, encoding="utf-8") as f:
            return json.load(f)
    return {}

def sauvegarder_docs_index(idx):
    os.makedirs(DOCS_DIR, exist_ok=True)
    with open(DOCS_INDEX, "w", encoding="utf-8") as f:
        json.dump(idx, f, ensure_ascii=False, indent=2)

def docs_de(email):
    return charger_docs_index().get(email, [])

def humaniser_taille(octets):
    for unite in ("o", "Ko", "Mo"):
        if octets < 1024 or unite == "Mo":
            return f"{octets:.0f} {unite}" if unite == "o" else f"{octets:.1f} {unite}"
        octets /= 1024


@app.route("/admin/employes", methods=["GET", "POST"])
def admin_employes():
    if not session.get("admin"):
        return redirect(url_for("admin"))

    employes = charger_employes()
    message = None

    if request.method == "POST":
        action = request.form.get("action")

        if action == "ajouter":
            prenom = request.form.get("prenom", "").strip()
            nom = request.form.get("nom", "").strip()
            email = request.form.get("email", "").strip()
            if prenom and nom and email:
                employes.append({"prenom": prenom, "nom": nom, "email": email})
                sauvegarder_employes(employes)
                message = f"{prenom} {nom} ajouté."

        elif action == "supprimer":
            email = request.form.get("email", "")
            employes = [e for e in employes if e["email"] != email]
            sauvegarder_employes(employes)
            message = "Employé supprimé."

        elif action == "modifier":
            email_orig = request.form.get("email_orig", "")
            for e in employes:
                if e["email"] == email_orig:
                    e["prenom"] = request.form.get("prenom", e["prenom"]).strip()
                    e["nom"] = request.form.get("nom", e["nom"]).strip()
                    e["email"] = request.form.get("email", e["email"]).strip()
                    break
            sauvegarder_employes(employes)
            message = "Employé modifié."

        employes = charger_employes()

    # Enrichit chaque employé (poste, nb docs, alertes, documents manquants, statut)
    profils = charger_profils()
    idx_docs = charger_docs_index()
    actifs, archives = [], []
    for e in employes:
        profil = profils.get(e["email"], {})
        info = {
            **e,
            "poste": profil.get("poste", ""),
            "contrat": profil.get("type_contrat", ""),
            "nb_docs": len(idx_docs.get(e["email"], [])),
            "alertes": alertes_completes(e["email"], profil),
            "manquants": docs_manquants(e["email"]),
            "statut": profil.get("statut", "actif"),
            "photo": profil.get("photo"),
        }
        (archives if info["statut"] == "archive" else actifs).append(info)

    postes = sorted({i["poste"] for i in actifs if i["poste"]})
    contrats = sorted({i["contrat"] for i in actifs if i["contrat"]})

    # Synthèse RH (intégrée ici) : actifs ayant une alerte ou un document manquant
    synthese = [i for i in actifs if i["alertes"] or i["manquants"]]
    synthese.sort(key=lambda i: (any(a["niveau"] == "rouge" for a in i["alertes"]),
                                 len(i["alertes"]), len(i["manquants"])), reverse=True)
    nb_rouge = sum(1 for i in synthese for a in i["alertes"] if a["niveau"] == "rouge")
    nb_manquants = sum(len(i["manquants"]) for i in synthese)

    return render_template("admin_employes.html", employes=actifs, archives=archives,
                           message=message, postes=postes, contrats=contrats,
                           synthese=synthese, nb_rouge=nb_rouge, nb_manquants=nb_manquants)


@app.route("/mon-espace")
def mon_espace():
    token = request.args.get("token", "")
    prenom = request.args.get("prenom", "")
    if not token or not prenom:
        abort(404)

    emp = resoudre_employe(token, charger_employes())
    token_canon = generer_token(emp["prenom"], emp["email"]) if emp else token

    historique = []
    for fichier in sorted(os.listdir(BASE_DIR), reverse=True):
        if fichier.startswith("reponses_") and fichier.endswith(".json"):
            parts = fichier.replace("reponses_","").replace(".json","").split("_")
            if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                m, a = int(parts[0]), int(parts[1])
                with open(os.path.join(BASE_DIR, fichier), encoding="utf-8") as fp:
                    reponses = json.load(fp)
                r = reponse_de(reponses, emp["prenom"], emp["email"]) if emp else reponses.get(token)
                if r:
                    solde = round(r["heures_plus"] - r["heures_moins"], 2)
                    historique.append({
                        "mois_annee": f"{MOIS_FR[m]} {a}",
                        "heures_plus": r["heures_plus"],
                        "heures_moins": r["heures_moins"],
                        "solde": solde,
                        "commentaire": r.get("commentaire", ""),
                        "date": r["date"],
                        "lien": f"/releve?token={token}&prenom={prenom}",
                    })

    return render_template("mon_espace.html", prenom=prenom, token=token, historique=historique)


@app.route("/admin/absences")
def admin_absences():
    # Fusionné dans le tableau de bord (section Absentéisme).
    if not session.get("admin"):
        return redirect(url_for("admin"))
    return redirect(url_for("admin_dashboard"))


def _admin_absences_legacy():
    if not session.get("admin"):
        return redirect(url_for("admin"))

    employes = charger_employes()
    annee_courante = datetime.now().year
    annee = int(request.args.get("annee", annee_courante))
    mois_actuel = datetime.now().month if annee == annee_courante else 12

    annees_dispo = set()
    for fichier in os.listdir(BASE_DIR):
        if fichier.startswith("reponses_") and fichier.endswith(".json"):
            parts = fichier.replace("reponses_","").replace(".json","").split("_")
            if len(parts) == 2 and parts[1].isdigit():
                annees_dispo.add(int(parts[1]))
    annees_dispo.add(annee_courante)
    annees_dispo = sorted(annees_dispo, reverse=True)

    mois_disponibles = []
    stats = {}  # {prenom: {mois: h_moins}}

    for emp in employes:
        stats[emp["prenom"]] = {}

    for m in range(1, mois_actuel + 1):
        f = reponses_file(m, annee)
        if os.path.exists(f):
            with open(f, encoding="utf-8") as fp:
                reponses = json.load(fp)
            mois_disponibles.append(m)
            for emp in employes:
                token = generer_token(emp["prenom"], emp["email"])
                r = reponse_de(reponses, emp["prenom"], emp["email"])
                stats[emp["prenom"]][m] = round(r["heures_moins"], 2) if r else None

    # Calcul cumul H- et classement
    classement = []
    for emp in employes:
        p = emp["prenom"]
        total_moins = sum(stats[p][m] for m in mois_disponibles if stats[p].get(m) is not None)
        mois_max = None
        val_max = 0
        for m in mois_disponibles:
            v = stats[p].get(m)
            if v and v > val_max:
                val_max = v
                mois_max = m
        classement.append({
            "prenom": p,
            "nom": emp["nom"],
            "total": round(total_moins, 2),
            "mois_max": MOIS_FR[mois_max] if mois_max else "-",
            "val_max": val_max,
            "mois_data": [stats[p].get(m, 0) or 0 for m in mois_disponibles],
        })

    classement.sort(key=lambda x: x["total"], reverse=True)

    return render_template("admin_absences.html",
        classement=classement,
        mois_disponibles=mois_disponibles,
        mois_noms={m: MOIS_FR[m][:3] for m in range(1, 13)},
        annee=annee,
        annees_dispo=annees_dispo,
    )


@app.route("/admin/dashboard")
def admin_dashboard():
    if not session.get("admin"):
        return redirect(url_for("admin"))

    employes = charger_employes()
    annee_courante = datetime.now().year
    annee = int(request.args.get("annee", annee_courante))
    mois_actuel = datetime.now().month if annee == annee_courante else 12

    # Trouver toutes les années disponibles
    annees_dispo = set()
    for fichier in os.listdir(BASE_DIR):
        if fichier.startswith("reponses_") and fichier.endswith(".json"):
            parts = fichier.replace("reponses_","").replace(".json","").split("_")
            if len(parts) == 2 and parts[1].isdigit():
                annees_dispo.add(int(parts[1]))
    annees_dispo.add(annee_courante)
    annees_dispo = sorted(annees_dispo, reverse=True)

    # Charger toutes les réponses de l'année
    donnees = {}  # {prenom: {mois: {h+, h-}}}
    mois_disponibles = []

    for m in range(1, mois_actuel + 1):
        f = reponses_file(m, annee)
        if os.path.exists(f):
            with open(f, encoding="utf-8") as fp:
                reponses = json.load(fp)
            mois_disponibles.append(m)
            for emp in employes:
                token = generer_token(emp["prenom"], emp["email"])
                r = reponse_de(reponses, emp["prenom"], emp["email"])
                if emp["prenom"] not in donnees:
                    donnees[emp["prenom"]] = {}
                if r:
                    donnees[emp["prenom"]][m] = {
                        "plus": r["heures_plus"],
                        "moins": r["heures_moins"],
                        "solde": round(r["heures_plus"] - r["heures_moins"], 2),
                        "commentaire": r.get("commentaire", ""),
                        "signature": r.get("signature", ""),
                        "date_signature": r.get("date_signature", ""),
                        "date": r.get("date", ""),
                        "jours": r.get("jours", []),
                    }
                else:
                    donnees[emp["prenom"]][m] = None

    # Calcul cumul annuel par employé
    cumuls = {}
    for emp in employes:
        p = emp["prenom"]
        total_plus = sum(donnees[p][m]["plus"] for m in mois_disponibles if donnees.get(p, {}).get(m))
        total_moins = sum(donnees[p][m]["moins"] for m in mois_disponibles if donnees.get(p, {}).get(m))
        cumuls[p] = {
            "plus": round(total_plus, 2),
            "moins": round(total_moins, 2),
            "solde": round(total_plus - total_moins, 2),
        }

    # Classement absentéisme
    classement_abs = []
    for emp in employes:
        p = emp["prenom"]
        total_moins = sum(donnees[p][m]["moins"] for m in mois_disponibles if donnees.get(p, {}).get(m))
        mois_max = None
        val_max = 0
        for m in mois_disponibles:
            d = donnees[p].get(m)
            if d and d["moins"] > val_max:
                val_max = d["moins"]
                mois_max = m
        classement_abs.append({
            "prenom": p,
            "nom": emp["nom"],
            "total": round(total_moins, 2),
            "mois_max": MOIS_FR[mois_max] if mois_max else "-",
            "val_max": val_max,
            "mois_data": [(donnees[p].get(m) or {}).get("moins", 0) or 0 for m in mois_disponibles],
        })
    classement_abs.sort(key=lambda x: x["total"], reverse=True)

    # Détail jour par jour, indexé par "prénom|mois" pour la fenêtre de détail.
    # Vide pour les relevés antérieurs à l'activation de cette sauvegarde.
    releves_jours = {}
    for emp in employes:
        p = emp["prenom"]
        for m in mois_disponibles:
            d = donnees[p].get(m)
            if d and d.get("jours"):
                releves_jours[f"{p}|{m}"] = d["jours"]

    return render_template("admin_dashboard.html",
        employes=employes,
        donnees=donnees,
        cumuls=cumuls,
        mois_disponibles=mois_disponibles,
        mois_noms={m: MOIS_FR[m][:3] for m in range(1, 13)},
        annee=annee,
        annees_dispo=annees_dispo,
        classement_abs=classement_abs,
        releves_jours=releves_jours,
    )


@app.route("/admin/historique")
def admin_historique():
    if not session.get("admin"):
        return redirect(url_for("admin"))

    employes = charger_employes()
    nb_total = len(employes)
    historique = []

    for fichier in sorted(os.listdir(BASE_DIR), reverse=True):
        if fichier.startswith("reponses_") and fichier.endswith(".json"):
            parts = fichier.replace("reponses_", "").replace(".json", "").split("_")
            if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                m, a = int(parts[0]), int(parts[1])
                with open(os.path.join(BASE_DIR, fichier), encoding="utf-8") as f:
                    data = json.load(f)
                historique.append({
                    "mois": m,
                    "annee": a,
                    "label": f"{MOIS_FR[m]} {a}",
                    "repondus": len(data),
                    "total": nb_total,
                    "fichier": f"{m}_{a}",
                })

    return render_template("admin_historique.html", historique=historique)


@app.route("/admin/historique/<int:mois>/<int:annee>")
def admin_historique_mois(mois, annee):
    if not session.get("admin"):
        return redirect(url_for("admin"))

    reponses = charger_reponses(mois, annee)
    employes = charger_employes()
    mois_annee = f"{MOIS_FR[mois]} {annee}"

    resultats = []
    for emp in employes:
        token = generer_token(emp["prenom"], emp["email"])
        reponse = reponse_de(reponses, emp["prenom"], emp["email"])
        resultats.append({
            "prenom": emp["prenom"],
            "nom": emp["nom"],
            "repondu": reponse is not None,
            "heures_plus": reponse["heures_plus"] if reponse else "-",
            "heures_moins": reponse["heures_moins"] if reponse else "-",
            "commentaire": reponse.get("commentaire", "") if reponse else "",
            "signature": reponse.get("signature", "") if reponse else "",
            "date": reponse["date"] if reponse else "-",
        })

    repondus = sum(1 for r in resultats if r["repondu"])
    return render_template("admin_historique_mois.html", resultats=resultats,
                           repondus=repondus, total=len(resultats),
                           mois_annee=mois_annee, mois=mois, annee=annee)


@app.route("/admin/historique/export/<int:mois>/<int:annee>")
def admin_historique_export(mois, annee):
    if not session.get("admin"):
        return redirect(url_for("admin"))

    reponses = charger_reponses(mois, annee)
    employes = charger_employes()
    mois_annee = f"{MOIS_FR[mois]} {annee}"

    wb = Workbook()
    ws = wb.active
    ws.title = f"Relevés {mois_annee}"
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:I1")
    ws["A1"] = f"Relevés d'heures — {mois_annee}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Prénom", "Nom", "H+", "H−", "Signature", "Date signature", "Commentaire", "Date envoi", "Statut"]
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E75B6")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for emp in employes:
        token = generer_token(emp["prenom"], emp["email"])
        r = reponse_de(reponses, emp["prenom"], emp["email"])
        if r:
            row = [emp["prenom"], emp["nom"], r["heures_plus"], r["heures_moins"],
                   r.get("signature", ""), r.get("date_signature", ""), r.get("commentaire", ""), r["date"], "Reçu"]
            fill = PatternFill("solid", fgColor="C6EFCE")
        else:
            row = [emp["prenom"], emp["nom"], "-", "-", "", "", "", "-", "En attente"]
            fill = PatternFill("solid", fgColor="FFEB9C")
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        ws[ws.max_row][8].fill = fill

    for i, w in enumerate([14, 16, 8, 8, 20, 14, 30, 16, 12], 1):
        ws.column_dimensions[chr(64+i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"Releves_{MOIS_FR[mois]}_{annee}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/admin/planning", methods=["GET", "POST"])
def admin_planning():
    if not session.get("admin"):
        return redirect(url_for("admin"))

    employes = charger_employes()
    planning = charger_planning()
    mois = datetime.now().month
    annee = datetime.now().year

    img_name = f"planning_{mois}_{annee}.png"
    img_path = os.path.join(UPLOAD_FOLDER, img_name)
    img_url = f"/static/planning_img/{img_name}" if os.path.exists(img_path) else None

    if request.method == "POST":
        # Sauvegarde image si uploadée
        if "image" in request.files:
            img = request.files["image"]
            if img and img.filename:
                img.save(img_path)
                img_url = f"/static/planning_img/{img_name}"

        nouveau = {}
        for emp in employes:
            p = emp["prenom"]
            trame = request.form.get(f"trame_{p}", "").strip()
            total = request.form.get(f"total_{p}", "").strip()
            absences = request.form.get(f"absences_{p}", "").strip()
            jours = request.form.get(f"jours_{p}", "").strip()
            nouveau[p] = {
                "trame": hm_to_float(trame) if trame else 0,
                "total": hm_to_float(total) if total else 0,
                "absences": absences,
                "jours": int(jours) if jours.isdigit() else 0,
            }
        sauvegarder_planning(nouveau)
        return redirect(url_for("admin"))

    return render_template("admin_planning.html", employes=employes, planning=planning,
                           mois_annee=f"{MOIS_FR[mois]} {annee}", img_url=img_url)


@app.route("/admin/planning/supprimer-image", methods=["POST"])
def admin_planning_supprimer_image():
    """Supprime l'image de planning chargée pour le mois en cours."""
    if not session.get("admin"):
        return redirect(url_for("admin"))
    mois, annee = datetime.now().month, datetime.now().year
    img_path = os.path.join(UPLOAD_FOLDER, f"planning_{mois}_{annee}.png")
    try:
        if os.path.exists(img_path):
            os.remove(img_path)
    except OSError:
        app.logger.exception("Échec suppression image planning")
    return redirect(url_for("admin_planning"))


@app.route("/admin/planning/reset", methods=["POST"])
def admin_planning_reset():
    """Réinitialise le remplissage du planning (les données saisies) du mois en cours."""
    if not session.get("admin"):
        return redirect(url_for("admin"))
    sauvegarder_planning({})
    return redirect(url_for("admin_planning"))


@app.route("/admin/logout")
def admin_logout():
    session.pop("admin", None)
    return redirect(url_for("admin"))


@app.route("/admin/erreurs")
def admin_erreurs():
    if not session.get("admin"):
        return redirect(url_for("admin"))
    import html as _html
    chemin = os.path.join(BASE_DIR, "logs", "erreurs.log")
    lignes = []
    if os.path.exists(chemin):
        with open(chemin, encoding="utf-8") as f:
            lignes = f.readlines()[-200:]
    contenu = _html.escape("".join(reversed(lignes))) or "Aucune erreur enregistrée. 🎉"
    return (f"<h2 style='font-family:Arial;color:#1F4E79'>Journal des erreurs (récentes en haut)</h2>"
            f"<p style='font-family:Arial'><a href='/admin'>← Retour admin</a></p>"
            f"<pre style='font-family:monospace;font-size:13px;background:#f5f5f5;"
            f"padding:16px;border-radius:8px;white-space:pre-wrap'>{contenu}</pre>")


@app.route("/admin/valider", methods=["POST"])
def admin_valider():
    """N — l'admin valide (ou annule) un relevé reçu, avant la paie."""
    if not session.get("admin"):
        return redirect(url_for("admin"))
    email = request.form.get("email", "")
    valide = request.form.get("valide") == "1"
    emp = next((e for e in charger_employes() if e["email"] == email), None)
    if emp:
        reponses = charger_reponses()
        for t in tokens_valides(emp["prenom"], emp["email"]):
            if t in reponses:
                reponses[t]["valide"] = valide
                reponses[t]["date_validation"] = (
                    datetime.now().strftime("%d/%m/%Y %H:%M") if valide else "")
                ecrire_reponses(reponses)
                break
    return redirect(url_for("admin"))


@app.route("/admin/employe")
def admin_employe():
    """I — fiche annuelle d'un employé : 12 mois (H+/H−/solde) + tendance."""
    if not session.get("admin"):
        return redirect(url_for("admin"))
    email = request.args.get("email", "")
    annee = int(request.args.get("annee", datetime.now().year))
    emp = next((e for e in charger_employes() if e["email"] == email), None)
    if not emp:
        abort(404)
    mois_data = []
    for m in range(1, 13):
        r = reponse_de(charger_reponses(m, annee), emp["prenom"], emp["email"])
        if r:
            mois_data.append({"mois": MOIS_FR[m][:3], "plus": r["heures_plus"],
                              "moins": r["heures_moins"],
                              "solde": round(r["heures_plus"] - r["heures_moins"], 2),
                              "rempli": True, "valide": r.get("valide", False),
                              "commentaire": r.get("commentaire", "")})
        else:
            mois_data.append({"mois": MOIS_FR[m][:3], "plus": None, "moins": None,
                              "solde": None, "rempli": False, "valide": False, "commentaire": ""})
    cp = sum(d["plus"] for d in mois_data if d["rempli"])
    cm = sum(d["moins"] for d in mois_data if d["rempli"])
    return render_template("admin_employe.html", emp=emp, annee=annee, mois_data=mois_data,
                           cumul_plus=round(cp, 2), cumul_moins=round(cm, 2),
                           cumul_solde=round(cp - cm, 2),
                           profil=profil_de(email), champs_profil=CHAMPS_PROFIL,
                           documents=docs_de(email),
                           docs_par_famille=grouper_docs_par_famille(docs_de(email)),
                           familles_docs=FAMILLES_DOCS,
                           doc_err=request.args.get("doc_err"),
                           alertes=alertes_completes(email, profil_de(email)),
                           manquants=docs_manquants(email),
                           docs_requis=DOCS_REQUIS,
                           statut=profil_de(email).get("statut", "actif"),
                           journal=journal_trie(profil_de(email).get("journal", [])),
                           types_evenement=TYPES_EVENEMENT,
                           a_photo=bool(profil_de(email).get("photo")),
                           taches_arrivee=TACHES_ARRIVEE, taches_depart=TACHES_DEPART,
                           check_arrivee=profil_de(email).get("check_arrivee", []),
                           check_depart=profil_de(email).get("check_depart", []))


@app.route("/admin/employe/document", methods=["POST"])
def admin_employe_document():
    """Dépose un document RH (privé admin) rattaché à un employé."""
    if not session.get("admin"):
        return redirect(url_for("admin"))
    email = request.form.get("email", "")
    emp = next((e for e in charger_employes() if e["email"] == email), None)
    if not emp:
        abort(404)
    f = request.files.get("fichier")
    if not f or not f.filename:
        return redirect(url_for("admin_employe", email=email, doc_err="vide"))
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in EXT_DOCS_OK:
        return redirect(url_for("admin_employe", email=email, doc_err="type"))
    os.makedirs(DOCS_DIR, exist_ok=True)
    doc_id = uuid.uuid4().hex[:12]
    stored = f"{doc_id}_{secure_filename(f.filename)}"
    chemin = os.path.join(DOCS_DIR, stored)
    f.save(chemin)
    idx = charger_docs_index()
    idx.setdefault(email, []).append({
        "id": doc_id,
        "fichier": stored,
        "nom_original": f.filename,
        "type": request.form.get("type", "Autre / divers"),
        "libelle": request.form.get("libelle", "").strip() or f.filename,
        "expiration": request.form.get("expiration", "").strip(),
        "taille": humaniser_taille(os.path.getsize(chemin)),
        "date_ajout": datetime.now().strftime("%d/%m/%Y %H:%M"),
    })
    sauvegarder_docs_index(idx)
    return redirect(url_for("admin_employe", email=email))


@app.route("/admin/document/<doc_id>")
def admin_document(doc_id):
    """Télécharge un document RH (admin uniquement)."""
    if not session.get("admin"):
        return redirect(url_for("admin"))
    for docs in charger_docs_index().values():
        for d in docs:
            if d["id"] == doc_id:
                chemin = os.path.join(DOCS_DIR, d["fichier"])
                if os.path.exists(chemin):
                    return send_file(chemin, as_attachment=True, download_name=d["nom_original"])
    abort(404)


@app.route("/admin/employe/document/supprimer", methods=["POST"])
def admin_employe_document_supprimer():
    """Supprime un document RH précis (action admin explicite)."""
    if not session.get("admin"):
        return redirect(url_for("admin"))
    email = request.form.get("email", "")
    doc_id = request.form.get("id", "")
    idx = charger_docs_index()
    docs = idx.get(email, [])
    cible = next((d for d in docs if d["id"] == doc_id), None)
    if cible:
        chemin = os.path.join(DOCS_DIR, cible["fichier"])
        try:
            if os.path.exists(chemin):
                os.remove(chemin)
        except OSError:
            app.logger.exception("Échec suppression fichier document RH")
        idx[email] = [d for d in docs if d["id"] != doc_id]
        sauvegarder_docs_index(idx)
    return redirect(url_for("admin_employe", email=email))


@app.route("/admin/employe/profil", methods=["POST"])
def admin_employe_profil():
    """Enregistre le dossier salarié (profil RH) d'un employé."""
    if not session.get("admin"):
        return redirect(url_for("admin"))
    email = request.form.get("email", "")
    emp = next((e for e in charger_employes() if e["email"] == email), None)
    if not emp:
        abort(404)
    profils = charger_profils()
    profil = profils.get(email, {})  # conserve les clés hors formulaire (ex: statut)
    for cle, _ in CHAMPS_PROFIL:
        profil[cle] = request.form.get(cle, "").strip()
    profils[email] = profil
    sauvegarder_profils(profils)
    return redirect(url_for("admin_employe", email=email, annee=request.form.get("annee", datetime.now().year)))


@app.route("/admin/employe/statut", methods=["POST"])
def admin_employe_statut():
    """Archive ou réactive un employé (statut conservé dans son profil)."""
    if not session.get("admin"):
        return redirect(url_for("admin"))
    email = request.form.get("email", "")
    statut = request.form.get("statut", "actif")
    emp = next((e for e in charger_employes() if e["email"] == email), None)
    if not emp:
        abort(404)
    profils = charger_profils()
    profil = profils.get(email, {})
    profil["statut"] = "archive" if statut == "archive" else "actif"
    profils[email] = profil
    sauvegarder_profils(profils)
    return redirect(request.form.get("retour") or url_for("admin_employes"))


@app.route("/admin/employe/export")
def admin_employe_export():
    """Exporte tout le dossier d'un employé (résumé + documents) en ZIP."""
    if not session.get("admin"):
        return redirect(url_for("admin"))
    import zipfile
    email = request.args.get("email", "")
    emp = next((e for e in charger_employes() if e["email"] == email), None)
    if not emp:
        abort(404)
    profil = profil_de(email)

    # Résumé texte du dossier
    lignes = [f"DOSSIER SALARIÉ — {emp['prenom']} {emp['nom']}", f"Email : {email}", ""]
    for cle, libelle in CHAMPS_PROFIL:
        lignes.append(f"{libelle} : {profil.get(cle, '') or '—'}")
    resume = "\r\n".join(lignes)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("dossier.txt", resume)
        for d in docs_de(email):
            chemin = os.path.join(DOCS_DIR, d["fichier"])
            if os.path.exists(chemin):
                arc = f"{d.get('type', 'Document')}/{d.get('nom_original', d['fichier'])}"
                zf.write(chemin, arcname=arc)
    buf.seek(0)
    nom = f"Dossier_{emp['prenom']}_{emp['nom']}.zip".replace(" ", "_")
    return send_file(buf, as_attachment=True, download_name=nom, mimetype="application/zip")


@app.route("/admin/employe/journal", methods=["POST"])
def admin_employe_journal_ajout():
    """Ajoute un événement au journal RH d'un employé."""
    if not session.get("admin"):
        return redirect(url_for("admin"))
    email = request.form.get("email", "")
    emp = next((e for e in charger_employes() if e["email"] == email), None)
    if not emp:
        abort(404)
    profils = charger_profils()
    profil = profils.get(email, {})
    journal = profil.get("journal", [])
    journal.append({
        "id": uuid.uuid4().hex[:8],
        "date": request.form.get("date", "").strip() or datetime.now().strftime("%d/%m/%Y"),
        "type": request.form.get("type", "Autre"),
        "note": request.form.get("note", "").strip(),
    })
    profil["journal"] = journal
    profils[email] = profil
    sauvegarder_profils(profils)
    return redirect(url_for("admin_employe", email=email))


@app.route("/admin/employe/journal/supprimer", methods=["POST"])
def admin_employe_journal_suppr():
    """Supprime un événement précis du journal RH."""
    if not session.get("admin"):
        return redirect(url_for("admin"))
    email = request.form.get("email", "")
    ev_id = request.form.get("id", "")
    profils = charger_profils()
    profil = profils.get(email, {})
    profil["journal"] = [e for e in profil.get("journal", []) if e.get("id") != ev_id]
    profils[email] = profil
    sauvegarder_profils(profils)
    return redirect(url_for("admin_employe", email=email))


@app.route("/admin/employe/photo", methods=["POST"])
def admin_employe_photo_upload():
    """Charge la photo de profil d'un employé."""
    if not session.get("admin"):
        return redirect(url_for("admin"))
    email = request.form.get("email", "")
    emp = next((e for e in charger_employes() if e["email"] == email), None)
    if not emp:
        abort(404)
    f = request.files.get("photo")
    if f and f.filename:
        ext = os.path.splitext(f.filename)[1].lower()
        if ext in {".png", ".jpg", ".jpeg", ".webp"}:
            os.makedirs(PHOTOS_DIR, exist_ok=True)
            nom = f"{uuid.uuid4().hex[:12]}{ext}"
            f.save(os.path.join(PHOTOS_DIR, nom))
            profils = charger_profils()
            profil = profils.get(email, {})
            ancienne = profil.get("photo")
            profil["photo"] = nom
            profils[email] = profil
            sauvegarder_profils(profils)
            if ancienne:  # supprime l'ancienne photo
                try:
                    os.remove(os.path.join(PHOTOS_DIR, ancienne))
                except OSError:
                    pass
    return redirect(url_for("admin_employe", email=email))


@app.route("/admin/employe/photo")
def admin_employe_photo_voir():
    """Sert la photo de profil (admin uniquement)."""
    if not session.get("admin"):
        return redirect(url_for("admin"))
    email = request.args.get("email", "")
    nom = profil_de(email).get("photo")
    if nom:
        chemin = os.path.join(PHOTOS_DIR, nom)
        if os.path.exists(chemin):
            return send_file(chemin)
    abort(404)


@app.route("/admin/employe/attestation")
def admin_employe_attestation():
    """Page imprimable d'attestation de travail, pré-remplie depuis le dossier."""
    if not session.get("admin"):
        return redirect(url_for("admin"))
    email = request.args.get("email", "")
    emp = next((e for e in charger_employes() if e["email"] == email), None)
    if not emp:
        abort(404)
    return render_template("attestation.html", emp=emp, profil=profil_de(email),
                           today=datetime.now().strftime("%d/%m/%Y"))


@app.route("/admin/employe/checklist", methods=["POST"])
def admin_employe_checklist():
    """Enregistre une checklist d'arrivée ou de départ d'un employé."""
    if not session.get("admin"):
        return redirect(url_for("admin"))
    email = request.form.get("email", "")
    typ = request.form.get("type")
    if typ not in ("arrivee", "depart"):
        abort(400)
    emp = next((e for e in charger_employes() if e["email"] == email), None)
    if not emp:
        abort(404)
    taches = TACHES_ARRIVEE if typ == "arrivee" else TACHES_DEPART
    cochees = [t for t in request.form.getlist("tache") if t in taches]
    profils = charger_profils()
    profil = profils.get(email, {})
    profil["check_" + typ] = cochees
    profils[email] = profil
    sauvegarder_profils(profils)
    return redirect(url_for("admin_employe", email=email))


@app.route("/admin/synthese")
def admin_synthese():
    # Intégrée dans la page Équipe (/admin/employes).
    if not session.get("admin"):
        return redirect(url_for("admin"))
    return redirect(url_for("admin_employes") + "#synthese")


def _admin_synthese_legacy():
    if not session.get("admin"):
        return redirect(url_for("admin"))
    profils = charger_profils()
    lignes = []
    for e in charger_employes():
        profil = profils.get(e["email"], {})
        if profil.get("statut") == "archive":
            continue
        al = alertes_completes(e["email"], profil)
        manq = docs_manquants(e["email"])
        if al or manq:
            lignes.append({
                "prenom": e["prenom"], "nom": e["nom"], "email": e["email"],
                "poste": profil.get("poste", ""),
                "alertes": al, "manquants": manq,
                "a_rouge": any(a["niveau"] == "rouge" for a in al),
            })
    lignes.sort(key=lambda l: (l["a_rouge"], len(l["alertes"]), len(l["manquants"])), reverse=True)
    return render_template("admin_synthese.html",
        lignes=lignes,
        nb_concernes=len(lignes),
        nb_alertes=sum(len(l["alertes"]) for l in lignes),
        nb_rouge=sum(1 for l in lignes for a in l["alertes"] if a["niveau"] == "rouge"),
        nb_manquants=sum(len(l["manquants"]) for l in lignes))


def construire_recap_xlsx(mois, annee):
    """Construit le classeur Excel récapitulatif des relevés du mois donné."""
    reponses = charger_reponses(mois, annee)
    employes = charger_employes()
    mois_annee = f"{MOIS_FR[mois]} {annee}"

    wb = Workbook()
    ws = wb.active
    ws.title = f"Relevés {mois_annee}"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:G1")
    ws["A1"] = f"Relevés d'heures — {mois_annee}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Prénom", "Nom", "H+", "H−", "Signature", "Date signature", "Commentaire", "Date envoi", "Statut"]
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E75B6")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for emp in employes:
        token = generer_token(emp["prenom"], emp["email"])
        r = reponse_de(reponses, emp["prenom"], emp["email"])
        if r:
            statut = "Reçu"
            fill = PatternFill("solid", fgColor="C6EFCE")
            row = [emp["prenom"], emp["nom"], r.get("heures_plus", "-"), r.get("heures_moins", "-"), r.get("signature", ""), r.get("date_signature", ""), r.get("commentaire", ""), r.get("date", "-"), statut]
        else:
            statut = "En attente"
            fill = PatternFill("solid", fgColor="FFEB9C")
            row = [emp["prenom"], emp["nom"], "-", "-", "", "", "", "-", statut]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        ws[ws.max_row][6].fill = fill

    for i, w in enumerate([14, 16, 8, 8, 20, 14, 30, 16, 12], 1):
        ws.column_dimensions[chr(64+i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@app.route("/admin/export")
def admin_export():
    if not session.get("admin"):
        return redirect(url_for("admin"))
    mois = datetime.now().month
    annee = datetime.now().year
    output = construire_recap_xlsx(mois, annee)
    return send_file(output, as_attachment=True,
                     download_name=f"Releves_{MOIS_FR[mois]}_{annee}.xlsx", mimetype=XLSX_MIME)


@app.route("/export_recap")
def export_recap():
    """Récap Excel des relevés du mois (pour l'envoi paie automatique par le
    runner GitHub). Clé requise."""
    cle = request.args.get("cle", "")
    if cle != API_CLE:
        abort(403)
    mois = int(request.args.get("mois", datetime.now().month))
    annee = int(request.args.get("annee", datetime.now().year))
    output = construire_recap_xlsx(mois, annee)
    return send_file(output, as_attachment=True,
                     download_name=f"Releves_{MOIS_FR[mois]}_{annee}.xlsx", mimetype=XLSX_MIME)


def envoyer_confirmation(sujet, message):
    import smtplib
    from email.mime.text import MIMEText
    from dotenv import load_dotenv
    load_dotenv(os.path.join(BASE_DIR, ".env"))
    gmail_user = os.getenv("GMAIL_USER")
    gmail_pwd = os.getenv("GMAIL_APP_PASSWORD")
    msg = MIMEText(message, "plain", "utf-8")
    msg["From"] = gmail_user
    msg["To"] = "pharmanantu@gmail.com"
    msg["Subject"] = sujet
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(gmail_user, gmail_pwd)
        server.sendmail(gmail_user, "pharmanantu@gmail.com", msg.as_string())


def notifier_releve(donnees):
    """Déclenche le workflow GitHub Actions 'nouveau_releve' qui génère le PDF
    du relevé et l'envoie à l'admin. Le serveur gratuit ne pouvant pas faire de
    SMTP, l'envoi est délégué au runner GitHub. Nécessite GITHUB_TOKEN dans le .env."""
    import urllib.request
    from dotenv import load_dotenv
    load_dotenv(os.path.join(BASE_DIR, ".env"))
    gh_token = os.getenv("GITHUB_TOKEN")
    if not gh_token:
        return
    url = "https://api.github.com/repos/pharmanantu-bit/botRh/dispatches"
    body = json.dumps({
        "event_type": "nouveau_releve",
        "client_payload": donnees,
    }).encode("utf-8")
    req = urllib.request.Request(url, data=body, method="POST")
    req.add_header("Authorization", f"Bearer {gh_token}")
    req.add_header("Accept", "application/vnd.github+json")
    req.add_header("Content-Type", "application/json")
    req.add_header("User-Agent", "botRh")
    urllib.request.urlopen(req, timeout=15)


@app.route("/deploy")
def deploy():
    import subprocess
    cle = request.args.get("cle", "")
    if cle != "botRh-deploy-2026":
        abort(403)
    try:
        result = subprocess.run(["git", "pull"], cwd="/home/pharmacie92000/botRh",
                                capture_output=True, text=True, timeout=30)
        # Toucher le fichier WSGI force le reload sur PythonAnywhere
        import pathlib
        pathlib.Path("/var/www/pharmacie92000_pythonanywhere_com_wsgi.py").touch()
        return f"OK\n{result.stdout}", 200
    except Exception as e:
        return f"Erreur: {e}", 500


@app.route("/trigger")
def trigger():
    cle = request.args.get("cle", "")
    if cle != API_CLE:
        abort(403)

    jour = datetime.now().day
    mois = datetime.now().month
    annee = datetime.now().year
    mois_annee = f"{MOIS_FR[mois]} {annee}"

    if jour == 20:
        from email_sender import send_emails, load_employees
        send_emails()
        nb = len(load_employees())
        envoyer_confirmation(
            f"botRh — Relevés {mois_annee} envoyés",
            f"Les relevés d'heures de {mois_annee} ont bien été envoyés à {nb} employés.\n\nConsulte l'admin : https://pharmacie92000.pythonanywhere.com/admin"
        )
        return "Envoi relevés OK", 200
    elif jour == 22:
        from relance_sender import send_relances, load_employees
        a_relancer = send_relances()
        nb_total = len(load_employees())
        nb_relances = len(a_relancer)
        if nb_relances == 0:
            msg_conf = f"Tous les employés ont répondu — aucune relance envoyée pour {mois_annee}."
        else:
            noms = ", ".join(e["prenom"] for e in a_relancer)
            msg_conf = f"{nb_relances}/{nb_total} relances envoyées pour {mois_annee}.\n\nEmployés relancés : {noms}\n\nConsulte l'admin : https://pharmacie92000.pythonanywhere.com/admin"
        envoyer_confirmation(f"botRh — Relances {mois_annee}", msg_conf)
        return "Envoi relances OK", 200
    else:
        return f"Rien à faire (jour {jour})", 200


@app.route("/export_reponses")
def export_reponses():
    """Renvoie les réponses (qui a rempli son relevé) du mois demandé, pour que
    les relances envoyées depuis GitHub Actions sachent qui relancer. Clé requise."""
    cle = request.args.get("cle", "")
    if cle != API_CLE:
        abort(403)
    mois = int(request.args.get("mois", datetime.now().month))
    annee = int(request.args.get("annee", datetime.now().year))
    f = reponses_file(mois, annee)
    data = {}
    if os.path.exists(f):
        with open(f, encoding="utf-8") as fp:
            data = json.load(fp)
    return app.response_class(json.dumps(data, ensure_ascii=False),
                              mimetype="application/json")


@app.route("/export_employes")
def export_employes():
    """Renvoie la liste des employés gérée via l'admin, pour que les envois
    (GitHub Actions) utilisent toujours la liste à jour. Le serveur est la
    source unique de vérité. Clé requise."""
    cle = request.args.get("cle", "")
    if cle != API_CLE:
        abort(403)
    return app.response_class(json.dumps(charger_employes(), ensure_ascii=False),
                              mimetype="application/json")


@app.route("/export_backup")
def export_backup():
    """Regroupe toutes les données (employés, réponses, planning) en un seul JSON
    pour la sauvegarde automatique quotidienne. Clé requise."""
    cle = request.args.get("cle", "")
    if cle != API_CLE:
        abort(403)
    data = {
        "genere_le": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "employes": charger_employes(),
        "reponses": {},
        "planning": {},
    }
    for fichier in os.listdir(BASE_DIR):
        if fichier.endswith(".json") and (fichier.startswith("reponses_") or fichier.startswith("planning_")):
            try:
                with open(os.path.join(BASE_DIR, fichier), encoding="utf-8") as f:
                    cible = "reponses" if fichier.startswith("reponses_") else "planning"
                    data[cible][fichier] = json.load(f)
            except Exception:
                app.logger.exception(f"Sauvegarde : lecture de {fichier} impossible")
    return app.response_class(json.dumps(data, ensure_ascii=False, indent=2),
                              mimetype="application/json")


@app.route("/healthcheck")
def healthcheck():
    """Diagnostic de l'état du serveur (config .env, employés, documents).
    Ne renvoie aucune valeur secrète, seulement des booléens/compteurs.
    Avec &smtp=1 : teste l'authentification Gmail sans envoyer d'e-mail."""
    import json as _json
    from dotenv import load_dotenv
    load_dotenv(os.path.join(BASE_DIR, ".env"))
    cle = request.args.get("cle", "")
    if cle != API_CLE:
        abort(403)

    gmail_user = os.getenv("GMAIL_USER")
    gmail_pwd = os.getenv("GMAIL_APP_PASSWORD")
    info = {
        "gmail_user_set": bool(gmail_user),
        "gmail_pwd_set": bool(gmail_pwd) and not str(gmail_pwd).startswith("x"),
        "smtp_login_ok": None,
    }

    try:
        from email_sender import load_employees
        info["employees_count"] = len(load_employees())
    except Exception as e:
        info["employees_error"] = str(e)

    docs_dir = os.path.join(BASE_DIR, "documents")
    info["documents_count"] = (
        len([f for f in os.listdir(docs_dir) if f.endswith(".docx")])
        if os.path.isdir(docs_dir) else 0
    )

    info["github_token_set"] = bool(os.getenv("GITHUB_TOKEN"))
    if request.args.get("github") == "1":
        import urllib.request
        try:
            req = urllib.request.Request("https://api.github.com/", headers={"User-Agent": "botRh"})
            with urllib.request.urlopen(req, timeout=15) as r:
                info["github_reachable"] = (r.status == 200)
        except Exception as e:
            info["github_reachable"] = False
            info["github_error"] = str(e)

    if request.args.get("dispatch") == "1":
        import urllib.request, urllib.error
        gh_token = os.getenv("GITHUB_TOKEN")
        try:
            corps = json.dumps({
                "event_type": "nouveau_releve",
                "client_payload": {"prenom": "DIAG", "heures_plus": 0, "heures_moins": 0,
                                   "commentaire": "diagnostic dispatch", "date_signature": "",
                                   "signature": "", "date": "", "mois": datetime.now().month,
                                   "annee": datetime.now().year},
            }).encode("utf-8")
            req = urllib.request.Request(
                "https://api.github.com/repos/pharmanantu-bit/botRh/dispatches",
                data=corps, method="POST")
            req.add_header("Authorization", f"Bearer {gh_token}")
            req.add_header("Accept", "application/vnd.github+json")
            req.add_header("Content-Type", "application/json")
            req.add_header("User-Agent", "botRh")
            with urllib.request.urlopen(req, timeout=15) as r:
                info["dispatch_status"] = r.status
        except urllib.error.HTTPError as e:
            info["dispatch_status"] = e.code
            info["dispatch_error"] = e.read().decode("utf-8", "ignore")[:300]
        except Exception as e:
            info["dispatch_error"] = str(e)

    if request.args.get("smtp") == "1" and info["gmail_user_set"] and info["gmail_pwd_set"]:
        import smtplib
        try:
            s = smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=15)
            s.login(gmail_user, gmail_pwd)
            s.quit()
            info["smtp_login_ok"] = True
        except Exception as e:
            info["smtp_login_ok"] = False
            info["smtp_error"] = str(e)

    return app.response_class(_json.dumps(info, ensure_ascii=False, indent=2),
                              mimetype="application/json")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)
